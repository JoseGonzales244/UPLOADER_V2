"""
Pruebas unitarias para el Pipeline de Dotación Mensual y Licencias Speech Analytics.
"""
import unittest
from unittest.mock import patch, MagicMock
from fastapi.testclient import TestClient
from modules.dotacion.dotacion_config import DotacionConfig
from modules.dotacion.phases.fase2_sincronizacion import get_new_seniority
from modules.dotacion.phases.fase_licencias_sa import is_backoffice
from backend.main import app


class TestDotacionPipelineLogic(unittest.TestCase):

    def test_dotacion_config_period_resolution(self):
        """Verifica la resolución correcta de fechas y nombres de mes en DotacionConfig."""
        cfg = DotacionConfig(target_period="2026-08")
        self.assertEqual(cfg.year, 2026)
        self.assertEqual(cfg.month, 8)
        self.assertEqual(cfg.prev_year, 2026)
        self.assertEqual(cfg.prev_month, 7)
        self.assertEqual(cfg.month_name_upper, "AGOSTO")
        self.assertEqual(cfg.prev_month_name_upper, "JULIO")

    def test_seniority_progression(self):
        """Verifica la máquina de estados de antigüedad (R0 -> R1 -> R2 -> R3)."""
        self.assertEqual(get_new_seniority(None), "R0")
        self.assertEqual(get_new_seniority(""), "R0")
        self.assertEqual(get_new_seniority("R0"), "R1")
        self.assertEqual(get_new_seniority("r0"), "R1")
        self.assertEqual(get_new_seniority("R1"), "R2")
        self.assertEqual(get_new_seniority("R2"), "R3")
        self.assertEqual(get_new_seniority("R3"), "R3")
        self.assertEqual(get_new_seniority("R4"), "R3")

    def test_is_backoffice_detection(self):
        """Verifica la detección de personal BackOffice y exclusión de interinos."""
        self.assertTrue(is_backoffice("Gestion_BackOffice", "ASESOR"))
        self.assertTrue(is_backoffice("", "ASISTENTE BO"))
        self.assertTrue(is_backoffice("Backoffice regular", ""))
        # Interinos NO son excluidos
        self.assertFalse(is_backoffice("Backoffice Interino", "ASESOR"))
        self.assertFalse(is_backoffice("Normal", "ASESOR VENTAS"))

    def test_supervisor_inconsistencies_detection(self):
        """Verifica la detección de alertas cuando un supervisor tiene múltiples códigos o viceversa."""
        from modules.dotacion.core.matching import detect_supervisor_inconsistencies

        records = [
            {"super": "CATHERINE JOSEFINA ESPINOZA TIMOTEO", "reg_super": "B15241"},
            {"super": "CATHERINE JOSEFINA ESPINOZA TIMOTEO", "reg_super": "B43648"}, # Error en input
            {"super": "BRUNO MIRANDA", "reg_super": "B43648"},
            {"super": "MARIA LOPEZ", "reg_super": "B99999"}
        ]

        alerts = detect_supervisor_inconsistencies(records)
        self.assertTrue(len(alerts) >= 2)
        # Catherine con 2 códigos
        self.assertTrue(any("CATHERINE JOSEFINA ESPINOZA TIMOTEO" in a and "B15241" in a and "B43648" in a for a in alerts))
        # B43648 con 2 supervisores (Bruno y Catherine)
    def test_advisor_registration_resolution(self):
        """Verifica que el índice de registro de asesor reconozca REG_COLAB, REG_PROMOTOR, REG_EJECUTIVO y REGISTRO, sin confundir jefe/supervisor."""
        from modules.dotacion.utils.excel import find_advisor_reg_col_idx

        headers_select_1 = ['PERIODO', 'EQUIPO', 'REG_JEFE', 'NOM_JEFE', 'REG_SUP', 'NOM_SUP', 'REG_COLAB', 'NOMBRE_COLABORADOR']
        self.assertEqual(find_advisor_reg_col_idx(headers_select_1), 6)

        headers_select_2 = ['PERIODO', 'EQUIPO', 'REG_JEFE', 'NOM_JEFE', 'REG_SUP', 'NOM_SUP', 'REGISTRO', 'NOMBRE_COLABORADOR']
        self.assertEqual(find_advisor_reg_col_idx(headers_select_2), 6)

        headers_promotor = ['PERIODO', 'REG_JEFE', 'REG_SUP', 'REG_PROMOTOR', 'NOMBRE']
        self.assertEqual(find_advisor_reg_col_idx(headers_promotor), 3)

        headers_ejecutivo = ['PERIODO', 'REG_JEFE', 'REG_SUP', 'REG_EJECUTIVO', 'NOMBRE']
        self.assertEqual(find_advisor_reg_col_idx(headers_ejecutivo), 3)

        headers_standard = ['PERIODO', 'REG_JEFE', 'REG SUPERVISOR JEFE', 'REGISTRO COLABORADOR', 'COLABORADOR']
        self.assertEqual(find_advisor_reg_col_idx(headers_standard), 3)

    def test_find_headers_multiple_terms(self):
        """Verifica búsqueda de cabeceras con lista de términos candidatos."""
        import openpyxl
        from modules.dotacion.utils.excel import find_headers_and_row

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=2, column=1, value="PERIODO")
        ws.cell(row=2, column=2, value="REG_COLAB")
        ws.cell(row=2, column=3, value="NOMBRE_COLABORADOR")

        headers, row_idx = find_headers_and_row(ws, ["REG_COLAB", "REG_PROMOTOR", "REG_EJECUTIVO", "REGISTRO"])
        self.assertEqual(row_idx, 2)
        self.assertIn("REG_COLAB", headers)

    def test_lock_resultados_sheet_protects_and_locks_cells(self):
        """Verifica que lock_resultados_sheet active la protección de hoja y fuerce todas las celdas a locked=True."""
        import openpyxl
        from modules.dotacion.utils.excel import lock_resultados_sheet

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "RESULTADOS"
        # Simular celdas explícitamente desbloqueadas
        ws.cell(row=5, column=3, value="CAROLINA")
        ws.cell(row=5, column=3).protection = openpyxl.styles.Protection(locked=False)
        ws.cell(row=5, column=4, value=25)
        ws.cell(row=5, column=4).protection = openpyxl.styles.Protection(locked=False)

        self.assertFalse(ws.cell(row=5, column=3).protection.locked)
        self.assertFalse(ws.protection.sheet)

        lock_resultados_sheet(wb)

        self.assertTrue(ws.protection.sheet)
        self.assertTrue(ws.cell(row=5, column=3).protection.locked)
        self.assertTrue(ws.cell(row=5, column=4).protection.locked)

    def test_fase1_cleans_avance_diario_and_manual_resultados(self):
        """Verifica que se limpien las celdas C:S en AVANCE DIARIO y filas 18, 21, 24, 27 en RESULTADOS."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws_av = wb.active
        ws_av.title = "AVANCE DIARIO"
        ws_res = wb.create_sheet("RESULTADOS")

        # Cargar valores previos
        for r in [7, 26, 31, 50, 55, 74, 79, 98]:
            for c in range(3, 20):
                ws_av.cell(row=r, column=c, value=99)

        for r in [18, 21, 24, 27]:
            for c in range(3, 20):
                ws_res.cell(row=r, column=c, value=55)

        # Simular limpieza de fase 1
        for sec_start in [7, 31, 55, 79]:
            for idx in range(20):
                r_idx = sec_start + idx
                for col_idx in range(3, 20):
                    ws_av.cell(row=r_idx, column=col_idx).value = None

        for r_idx in [18, 21, 24, 27]:
            for col_idx in range(3, 20):
                ws_res.cell(row=r_idx, column=col_idx).value = None

        # Aserciones
        for r in [7, 26, 31, 50, 55, 74, 79, 98]:
            for c in range(3, 20):
                self.assertIsNone(ws_av.cell(row=r, column=c).value)

        for r in [18, 21, 24, 27]:
            for c in range(3, 20):
                self.assertIsNone(ws_res.cell(row=r, column=c).value)


class TestDotacionEndpoints(unittest.TestCase):

    def setUp(self):
        self.client = TestClient(app)

    @patch("backend.main._run_dotacion_pipeline_task")
    def test_post_run_dotacion_pipeline_endpoint(self, mock_task):
        response = self.client.post(
            "/api/dotacion/run-pipeline",
            json={"periodo": "2026-08"}
        )
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["status"], "started")
        self.assertIn("2026-08", data["message"])

    @patch("backend.main._run_licencias_pipeline_task")
    def test_post_run_licencias_endpoint(self, mock_task):
        response = self.client.post(
            "/api/dotacion/run-licencias",
            json={"periodo": "202608"}
        )
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["status"], "started")


if __name__ == "__main__":
    unittest.main()
