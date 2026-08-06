import io
import re
import pandas as pd
import polars as pl

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


def _normalize_header(value) -> str:
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]+", " ", str(value).strip().lower()).strip()


def _looks_like_expected_header(row_values, expected_columns) -> bool:
    if not expected_columns:
        return False

    normalized_values = [_normalize_header(value) for value in row_values if value is not None]
    normalized_text = " ".join(normalized_values)
    matches = sum(1 for expected in expected_columns if _normalize_header(expected) in normalized_text)
    return matches >= 3


def _should_use_manual_excel_reader(selected_template: str | None) -> bool:
    if not selected_template:
        return False

    return (
        selected_template == "P001-CALIDAD_SA"
        or selected_template == "P030-RETENCION_CONVENIOS"
        or selected_template == "P025-SA_TCAD"
    )


def _read_excel_with_openpyxl(uploaded_file) -> pl.DataFrame | None:
    """Read the workbook manually for the P001-CALIDAD_SA template so the real header row is preserved."""
    if load_workbook is None:
        return None

    try:
        if isinstance(uploaded_file, str):
            workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        else:
            workbook = load_workbook(io.BytesIO(uploaded_file.getvalue()), read_only=True, data_only=True)
    except Exception:
        return None

    sheet_name = "Speech" if "Speech" in workbook.sheetnames else workbook.sheetnames[0]
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 29:
        return None

    header_row = rows[28]
    header_names = []
    for pos, value in enumerate(header_row):
        base_name = str(value).strip() if value is not None and str(value).strip() else f"column_{pos + 1}"
        header_names.append(base_name)

    data_rows = []
    for row in rows[29:]:
        values = list(row)
        if len(values) < len(header_names):
            values = values + [None] * (len(header_names) - len(values))
        elif len(values) > len(header_names):
            values = values[:len(header_names)]

        normalized_values = []
        for value in values:
            if value is None or (isinstance(value, str) and not value.strip()):
                normalized_values.append(None)
            else:
                normalized_values.append(value)
        data_rows.append(normalized_values)

    if not data_rows:
        return pl.DataFrame(schema=header_names)

    return pl.DataFrame(data_rows, schema=header_names, orient="row", infer_schema_length=None)


def read_excel_file(uploaded_file, selected_template: str | None = None, templates: dict | None = None) -> pl.DataFrame:
    """Reads Excel files (.xlsx, .xls) using the high-performance calamine engine or openpyxl for templates with header on row 29."""
    if _should_use_manual_excel_reader(selected_template):
        source = uploaded_file if isinstance(uploaded_file, str) else (uploaded_file.getvalue() if hasattr(uploaded_file, "getvalue") else uploaded_file)
        try:
            return pl.read_excel(source, engine="calamine", read_options={"header_row": 28})
        except Exception:
            df = _read_excel_with_openpyxl(uploaded_file)
            if df is not None:
                return df

    if isinstance(uploaded_file, str):
        df = pl.read_excel(uploaded_file, engine="calamine")
    else:
        df = pl.read_excel(uploaded_file.getvalue(), engine="calamine")
    return df


def read_csv_file(uploaded_file) -> pl.DataFrame:
    """Reads CSV files using Polars' fast multithreaded engine."""
    if isinstance(uploaded_file, str):
        return pl.read_csv(uploaded_file, truncate_ragged_lines=True)
    return pl.read_csv(uploaded_file.getvalue(), truncate_ragged_lines=True)

def read_unicode_text_file(uploaded_file) -> pl.DataFrame:
    """Reads Tab-separated Unicode text files using Polars."""
    if isinstance(uploaded_file, str):
        return pl.read_csv(uploaded_file, separator='\t', truncate_ragged_lines=True)
    return pl.read_csv(uploaded_file.getvalue(), separator='\t', truncate_ragged_lines=True)
