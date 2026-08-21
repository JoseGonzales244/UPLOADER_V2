import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

headers = [
    "Nombre de la información, base, reporte o dashboard",
    "Tipo de elemento",
    "Descripción breve / ¿qué contiene?",
    "¿Para qué se utiliza?",
    "KPI o indicador relacionado",
    "Rol en el flujo",
    "Frecuencia de actualización / uso",
    "Forma de actualización",
    "Tipo de fuente",
    "Fuente / repositorio conocido",
    "Nombre de tabla o archivo, si lo conoce",
    "Aplicativo / herramienta de uso",
    "Volumen conocido (GB)",
    "Quién utiliza la información",
    "Principal dificultad",
    "Impacto si falla",
    "Mejora esperada"
]

rows = [
    [
        "Funnel Piloto TCAD (Tarjetas Adicionales)",
        "Tabla / Reporte",
        "Información de las ventas cross de TC adicionales cruzadas con marcaciones de Speech Analytics y ventas reales.",
        "Mapeo del funnel de conversión del piloto TCAD, efectividad de ofertas y adopción del speech.",
        "Conversión ventas / Adopción TCAD",
        "Entrada / Analítica",
        "2 veces por semana / Mensual",
        "Semiautomática",
        "Tabla",
        "TDT (Teradata) / Speech Analytics",
        "DLAB_GEC.V_FNL_TCAD_SIMPLE, DLAB_GEC.M_EXP_CROSS_TCAD, DLAB_GEC.M_EXP_DATA_TCAD_SA, DLAB_GEC.V_EXP_VENTAS_TC_TCAD, E_DW_VIEWS.V_FCT_RT_TC_HISTORICO",
        "Power BI / Teradata SQL",
        "",
        "Supervisores, Jefaturas Comerciales y Analistas de Producto",
        "Adopción del speech y formato de reportes Verint",
        "Bajo",
        "Automatización total e integración al pipeline regular de TC"
    ],
    [
        "PBI Base Consumo (Colocaciones y Consentimiento)",
        "Dashboard / Base Consolidada",
        "Consolidado de colocaciones comerciales por producto (TC, PP, EC, CD, CON, IL, UPG, PA, SEG) con validación de consentimientos y líneas >40K.",
        "Seguimiento diario/quincenal del avance comercial de Televentas y alimentación de modelos de control.",
        "Ventas Totales / Colocaciones por Producto",
        "Fuente principal",
        "Diaria / Quincenal / Cierre",
        "Semiautomática",
        "Tabla / Vistas",
        "TDT (Teradata) / SQL Server / SharePoint",
        "DLAB_GEC.M_EXP_VENTAS_TC, M_EXP_VENTAS_PP, M_EXP_VENTAS_CD, M_EXP_VENTAS_EC, M_EXP_VENTAS_CON, M_EXP_VENTAS_IL, M_EXP_VENTAS_UPG, M_EXP_VENTAS_PA, M_EXP_VENTAS_SEG, DLAB_GEC.T_SP_CD40K, DLAB_GEC.BN_DESEMBOLSOS_GENERAL, DLAB_GEC.M_EXP_CONSUMO_SELECT_TC_CD_SEG, E_DW_VIEWS.V_FCT_RT_TC_HISTORICO",
        "Power BI / Teradata SQL / FastAPI",
        "",
        "Gerencia de Canales, Zonales, Jefaturas y Analistas",
        "Cruces multicanal y dependencias de sincronización de archivos auxiliares",
        "Alto",
        "Integración automatizada directa desde Data Warehouse sin intervención manual"
    ],
    [
        "PBI Evaluaciones Calidad (Consolidado Calidad NTD)",
        "Dashboard / Reporte Final",
        "Consolidado semanal y mensual de notas de calidad de asesores combinando evaluación manual de Insight con Speech Analytics de Verint, curvas y descuentos NTD.",
        "Monitoreo continuo de la calidad del servicio en llamadas de televentas, identificación de desvíos operativos y feedback formativo.",
        "% Nota Calidad Televentas / Tasa Errores Críticos (NTD)",
        "Fuente principal",
        "Semanal / Cierre Mensual",
        "Semiautomática",
        "Tabla / Vistas",
        "TDT (Teradata) / Insight PureCloud / Verint WFO",
        "DLAB_GEC.V_EXP_CALIDAD_NOTA_FINAL, DLAB_GEC.M_EXP_CALIDAD_PURECLOUD_PRE, DLAB_GEC.M_EXP_CALIDAD_DATA_SPEECH_ANALYTICS, DLAB_GEC.M_EXP_NTD_OBSERVACIONES_PRE, DLAB_GEC.M_EXP_CALIDAD_HISTORICO_ERRORES, DLAB_GEC.M_EXP_NTD_REPORTING_HISTORICO",
        "Power BI (Workspace Canales y Servicio al Cliente) / Teradata SQL",
        "",
        "Gerencia, Jefaturas, Supervisores de Calidad y Analistas",
        "Integración de fuentes heterogéneas (Insight + Verint SA + SharePoint)",
        "Crítico",
        "Gobernanza completa con ingesta directa vía API sin scrapers ni descargas intermedias"
    ],
    [
        "Dashboard KRI Ventas Sin Audio",
        "Dashboard / Reporte KRI",
        "Detección de operaciones comerciales colocadas por televentas que no cuentan con archivo de audio de respaldo o consentimiento grabado.",
        "Auditoría de riesgo operativo, control de cumplimiento normativo y mitigación de reclamos comerciales.",
        "% KRI Ventas Sin Audio",
        "Salida / KRI",
        "Diaria / Quincenal / Cierre",
        "Semiautomática",
        "Tabla",
        "TDT (Teradata) / Insight PureCloud",
        "DLAB_GEC.T_EXP_KRI_VENTAS_SINAUDIO, DLAB_GEC.T_EXP_KRI_VENTAS_SINAUDIO_CALIDAD, DLAB_GEC.M_EXP_CO_KRI_VENTA_TOTAL, DLAB_GEC.M_EXP_TRAFICO_GENESIS",
        "Power BI / Teradata SQL",
        "",
        "Gerencia de Canales, Riesgo Operacional, Compliance y Supervisores",
        "Cruce temporal entre grabaciones de tráfico telefónico y fecha efectiva de desembolso",
        "Crítico",
        "Trazabilidad nativa entre el identificador de llamada en el marcador y la venta en DW"
    ],
    [
        "Dashboard KRI Teléfonos No Autorizados",
        "Dashboard / Reporte KRI",
        "Identificación de llamadas y colocaciones realizadas a números telefónicos no registrados ni autorizados en la base de datos de clientes.",
        "Mitigación de riesgos legales y regulatorios frente a multas por infracción a la Ley de Protección de Datos Personales (LPDP).",
        "% KRI Teléfonos No Autorizados",
        "Salida / KRI",
        "Diaria / Quincenal / Cierre",
        "Semiautomática",
        "Tabla",
        "TDT (Teradata)",
        "DLAB_GEC.T_EXP_KRI_TELF_NO_AUTORIZADO, DLAB_GEC.T_EXP_KRI_TELF_NO_AUTORIZADO_CALIDAD, V_CONT_TELEFONO_APICLIENTE, TLV_CARGA_ACTUAL, TLV_CARGA_ACTUAL_DIGITAL",
        "Power BI / Teradata SQL",
        "",
        "Compliance, Riesgo Operacional, Calidad y Gerencia",
        "Heterogeneidad en formatos de teléfonos de contacto y actualización de la base cliente",
        "Crítico",
        "Validación automática previa al marcado telefónico en las listas de emisión"
    ],
    [
        "Consolidado Gerencial Cierre Mensual (Calidad & KRI)",
        "Base Consolidada / Reporte",
        "Resumen mensual definitivo de notas de calidad homologadas con jerarquías organizacionales (Supervisor, Jefe, Equipo) y métricas KRI consolidadas.",
        "Cierre operativo mensual para cálculo de comisiones e incentivos, reportes a comité gerencial y auditorías.",
        "Nota Gerencial Calidad / KRI Total Cierre",
        "Salida / Cierre",
        "Mensual (Cierre de mes)",
        "Semiautomática",
        "Tabla",
        "TDT (Teradata)",
        "DLAB_GEC.M_EXP_CALIDAD_NOTAS_TOTAL_GERENCIAL, DLAB_GEC.M_KRI_RESUMEN_TOTAL, DLAB_GEC.M_EXP_TELEVENTAS_EJECUTIVOS_GROUPED, DLAB_GEC.V_EXP_CALIDAD_NOTA_FINAL",
        "Teradata SQL / Power BI / Excel Gerencial",
        "",
        "Gerencia de Canales, Jefaturas y RRHH (Comisiones)",
        "Cambios en jerarquías del personal activo durante el mes de corte",
        "Crítico",
        "Sincronización automatizada de la matriz de personal con el sistema central de RRHH"
    ],
    [
        "Dashboard Piloto No Venta (Objeciones y Fuga de Ventas)",
        "Dashboard / Reporte Analítico",
        "Categorización automática por Speech Analytics de llamadas sin colocación (motivos: tasa, competencia, lo pensará, falta de rebatimiento) contrastada con ventas reales.",
        "Identificación de fugas comerciales, detección de ventas rescatadas y corrección de tipificaciones erróneas de asesores.",
        "% Fuga de Venta / % Rechazo por Tasa / Tasa de Rescate",
        "Fuente analítica / Piloto",
        "Mensual",
        "Semiautomática",
        "Tabla",
        "TDT (Teradata) / Speech Analytics",
        "DLAB_GEC.M_EXP_PILOTO_NO_VENTA, DLAB_GEC.M_EXP_STAGE_NO_VENTA, DLAB_GEC.M_EXP_VENTAS_TC, DLAB_GEC.M_EXP_VENTAS_PP, DLAB_GEC.M_EXP_VENTAS_EC, DLAB_GEC.M_EXP_VENTAS_CD, DLAB_GEC.M_EXP_VENTAS_SEG",
        "Power BI / Teradata SQL",
        "",
        "Jefaturas Comerciales, Supervisores y Analistas de Speech Analytics",
        "Procesamiento masivo de categorías de audio y descarte de falsos positivos en tipificación",
        "Medio",
        "Disponer de un tablero continuo de diagnóstico comercial en Power BI"
    ],
    [
        "Matriz Encuestas NPS Televentas",
        "Tabla / Base Analítica",
        "Tabla maestra de ventas y contactos del canal televentas estructurada por período, producto y ejecutivo para cruce con resultados de encuestas de experiencia.",
        "Medición de la satisfacción del cliente colocado por el canal televentas e identificación de factores de fricción.",
        "NPS Televentas / CSAT",
        "Fuente analítica",
        "Mensual",
        "Semiautomática",
        "Tabla",
        "TDT (Teradata)",
        "DLAB_GEC.F_NPS_VENTAS_TV, DLAB_GEC.M_EXP_VENTAS_TC, DLAB_GEC.M_EXP_VENTAS_PP, DLAB_GEC.M_EXP_VENTAS_EC, DLAB_GEC.M_EXP_VENTAS_CD, DLAB_GEC.M_EXP_VENTAS_SEG, DLAB_GEC.M_EXP_VENTAS_CON",
        "Power BI / Teradata SQL",
        "",
        "Analistas de Experiencia del Cliente (CX) y Gerencia",
        "Homologación de códigos de ejecutivo y estructura de productos por campaña",
        "Medio",
        "Carga automatizada mensual integrada al pipeline oficial de cierre"
    ],
    [
        "Repositorio y Solicitud de Audios Genesys",
        "Servicio / Archivo",
        "Repositorio local de grabaciones de llamadas (.mp3/.wav) y transcripciones descargadas a solicitud de correos de supervisores o formularios.",
        "Auditoría puntual de llamadas, atención de reclamos de clientes y validación de calidad en casos especiales.",
        "Tiempo de Atención de Solicitudes de Audio",
        "Servicio Operativo / Entrada",
        "Diaria / A demanda",
        "Semiautomática",
        "Archivos de Audio / Web CDP",
        "Genesys Cloud (PureCloud) / Outlook",
        "tracking.json, telefonos_cache.json, transcripciones_genesys/",
        "Google Chrome CDP / Python (Playwright, win32com)",
        "",
        "Supervisores, Analistas de Calidad y Equipo de Reclamos",
        "Dependencia de sesión interactiva de navegador abierta con depuración remota",
        "Medio",
        "Descarga desatendida mediante API oficial de grabaciones de Genesys Cloud"
    ]
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Reportes y Fuentes Finales"

header_fill = PatternFill(start_color="002A8F", end_color="002A8F", fill_type="solid")
header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style="thin", color="D3D3D3"),
    right=Side(style="thin", color="D3D3D3"),
    top=Side(style="thin", color="D3D3D3"),
    bottom=Side(style="thin", color="D3D3D3")
)

ws.append(headers)

for col_idx in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align
    cell.border = thin_border

zebra_fill = PatternFill(start_color="F7F9FC", end_color="F7F9FC", fill_type="solid")

for r_idx, row_data in enumerate(rows, start=2):
    ws.append(row_data)
    for c_idx in range(1, len(row_data) + 1):
        cell = ws.cell(row=r_idx, column=c_idx)
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = thin_border
        if r_idx % 2 == 0:
            cell.fill = zebra_fill

ws.row_dimensions[1].height = 35

col_widths = {
    1: 32, 2: 20, 3: 40, 4: 35, 5: 25, 6: 18, 7: 20, 8: 18,
    9: 16, 10: 25, 11: 45, 12: 25, 13: 14, 14: 28, 15: 35, 16: 14, 17: 38
}

for col_idx, width in col_widths.items():
    col_letter = get_column_letter(col_idx)
    ws.column_dimensions[col_letter].width = width

target_path = os.path.abspath("docs/INVENTARIO_FUENTES_APP_CALIDAD.xlsx")
os.makedirs(os.path.dirname(target_path), exist_ok=True)
wb.save(target_path)
print(f"Excel actualizado exitosamente en: {target_path}")
