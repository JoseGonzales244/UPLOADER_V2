"""
=============================================================================
EVALUADOR IA DE TRANSCRIPCIONES (PAGO AUTOMÁTICO TARJETAS DE CRÉDITO)
=============================================================================
Objetivo: 
  Lee las transcripciones extraídas en 'transcripciones_pa/' 
  y evalúa con Gemini si se ofreció y si el cliente aceptó/rechazó
  la afiliación al Pago Automático, actualizando 'Solicitud Cumplimiento TC 2026.xlsx'.
=============================================================================
"""
import os
import sys
import json
import time
import logging
from pathlib import Path
from typing import Optional, Dict, Any, List
import openpyxl
from dotenv import load_dotenv

PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

load_dotenv()

if sys.stdout and hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s", datefmt="%H:%M:%S")
logger = logging.getLogger("EvaluadorIAPA")

from infrastructure.llm.gemini_client import GeminiClient

TRANSCRIPTS_DIRS = [
    PROJECT_ROOT / "transcripciones_pa",
    PROJECT_ROOT / "data" / "transcripciones_pa"
]

EXCEL_FILE = PROJECT_ROOT / "Solicitud Cumplimiento TC 2026.xlsx"
AUDITED_FILE = PROJECT_ROOT / "Solicitud Cumplimiento TC 2026_Auditada.xlsx"


def find_transcript_file(dni_8: str, call_id: Optional[str]) -> Optional[Path]:
    """Busca el archivo .txt de transcripción para un DNI o CallID en las carpetas posibles."""
    for base_dir in TRANSCRIPTS_DIRS:
        if not base_dir.exists():
            continue

        if call_id and str(call_id).strip() not in ["", "None", "7464"]:
            direct_match = base_dir / f"TRANSCRIPT_DNI_{dni_8}_{call_id}.txt"
            if direct_match.exists():
                return direct_match

        # Búsqueda por patrón de DNI
        for f in base_dir.glob(f"TRANSCRIPT_DNI_{dni_8}_*.txt"):
            return f
        for f in base_dir.glob(f"TRANSCRIPT_DNI_*{dni_8}*.txt"):
            return f

        # Búsqueda por CallID
        if call_id and len(str(call_id)) > 10:
            for f in base_dir.glob(f"*{call_id}*.txt"):
                return f

    return None


def evaluate_text_with_gemini(full_text: str, llm_client: GeminiClient) -> Dict[str, Any]:
    prompt = f"""Eres un Auditor Senior de Cumplimiento de Televentas Bancarias de Interbank.
Tu ÚNICO objetivo es auditar la llamada para verificar si el asesor ofreció la AFILIACIÓN AL PAGO AUTOMÁTICO / DÉBITO AUTOMÁTICO de la Tarjeta de Crédito, y si el cliente ACEPTÓ o RECHAZÓ.

REGLAS DE EVALUACIÓN:
1. "NO_ACEPTA": El asesor ofreció afiliar a Pago/Débito Automático (o cargar el pago a su cuenta de ahorros), pero el cliente declinó, dijo que prefiere pagarlo por su cuenta/app, que no desea débito automático, etc.
   -> Debes capturar el timestamp exacto (mm:ss) del momento en que el CLIENTE rechaza o dice que no.
2. "ACEPTA": El asesor ofreció Pago Automático y el cliente dio su consentimiento explícito ("Sí", "De acuerdo", "Afílieme", "Claro").
   -> Debes capturar el timestamp exacto (mm:ss) del momento en que el CLIENTE acepta.
3. "NO_OFRECIDO": En toda la llamada NUNCA se mencionó la afiliación a Pago Automático ni Débito Automático.
   -> Timestamp: null.

TRANSCRIPCIÓN DE LA LLAMADA:
\"\"\"
{full_text}
\"\"\"

Responde estrictamente en formato JSON:
{{
  "estado": "NO_ACEPTA" | "ACEPTA" | "NO_OFRECIDO",
  "timestamp_cliente": "mm:ss" | null,
  "cita_textual_cliente": "Texto exacto donde el cliente responde o rechaza",
  "cita_textual_asesor": "Texto exacto donde el asesor ofrece el pago automático",
  "explicacion": "Breve justificación de 1 línea"
}}
"""
    try:
        response_str = llm_client.generate_content_with_retry(
            prompt=prompt,
            model_name="gemini-3.1-flash-lite",
            temperature=0.0,
            response_json=True
        )
        data = json.loads(response_str)
        estado = str(data.get("estado", "INCIERTO")).upper()
        ts = data.get("timestamp_cliente")

        if estado == "NO_ACEPTA":
            res_fmt = f"Cliente no acepta ({ts})" if ts else "Cliente no acepta"
        elif estado == "ACEPTA":
            res_fmt = f"Cliente acepta ({ts})" if ts else "Cliente acepta"
        elif estado == "NO_OFRECIDO":
            res_fmt = "No se ofreció Pago Automático"
        else:
            res_fmt = "REVISIÓN MANUAL PENDIENTE"

        return {
            "estado": estado,
            "timestamp": ts,
            "resultado_formateado": res_fmt,
            "cita": data.get("cita_textual_cliente", ""),
            "explicacion": data.get("explicacion", "")
        }
    except Exception as e:
        logger.error(f"Error evaluando con Gemini: {e}")
        return {
            "estado": "INCIERTO",
            "timestamp": None,
            "resultado_formateado": "REVISIÓN MANUAL PENDIENTE",
            "cita": "",
            "explicacion": str(e)
        }


def main():
    target_excel = AUDITED_FILE if AUDITED_FILE.exists() else EXCEL_FILE
    if not target_excel.exists():
        logger.critical(f"❌ No se encontró ningún archivo Excel ('{EXCEL_FILE}' ni '{AUDITED_FILE}')")
        return

    logger.info("=" * 75)
    logger.info("   INICIANDO EVALUACIÓN IA DE TRANSCRIPCIONES (PAGO AUTOMÁTICO TC)")
    logger.info(f"   Archivo Base : {target_excel.name}")
    logger.info("=" * 75)

    llm = GeminiClient(default_model="gemini-3.1-flash-lite")

    wb = openpyxl.load_workbook(target_excel)
    ws = wb.active

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    col_map = {str(h).strip().upper(): idx + 1 for idx, h in enumerate(headers) if h}
    col_dni = col_map.get("NRO DOCUMENTO", 4)
    col_res = col_map.get("RESULTADO", 7)
    col_id_llamada = col_map.get("ID LLAMADA", 8)
    col_fec_llamada = col_map.get("FECHA LLAMADA", 9)

    total_rows = ws.max_row
    evaluados = 0
    auditados_con_transcripcion = 0
    no_aceptados = 0
    aceptados = 0
    no_ofrecidos = 0
    pendientes_manual = 0

    for row_idx in range(2, total_rows + 1):
        dni_raw = ws.cell(row=row_idx, column=col_dni).value
        call_id_raw = ws.cell(row=row_idx, column=col_id_llamada).value

        if not dni_raw:
            continue

        dni_8 = str(int(dni_raw) if isinstance(dni_raw, float) else dni_raw).strip().zfill(8)
        call_id = str(call_id_raw).strip() if call_id_raw else ""

        matched_file = find_transcript_file(dni_8, call_id)

        evaluados += 1
        if matched_file and matched_file.exists():
            text_content = matched_file.read_text(encoding="utf-8")
            logger.info(f"\n--- [Caso {evaluados}/{total_rows - 1}] DNI: {dni_8} | Archivo: {matched_file.name} ---")
            eval_res = evaluate_text_with_gemini(text_content, llm)

            resultado_texto = eval_res["resultado_formateado"]
            estado = eval_res["estado"]

            if estado == "NO_ACEPTA":
                no_aceptados += 1
            elif estado == "ACEPTA":
                aceptados += 1
            elif estado == "NO_OFRECIDO":
                no_ofrecidos += 1
            else:
                pendientes_manual += 1

            logger.info(f"🎯 Dictamen : {resultado_texto}")
            logger.info(f"   Cita     : '{eval_res.get('cita')}'")
            logger.info(f"   Motivo   : {eval_res.get('explicacion')}")

            ws.cell(row=row_idx, column=col_res, value=resultado_texto)
            auditados_con_transcripcion += 1
            time.sleep(3)  # Pausa de cortesía para respetar Rate Limits de Gemini
        else:
            logger.warning(f"--- [Caso {evaluados}/{total_rows - 1}] DNI {dni_8}: Sin transcripción en carpetas ---")
            ws.cell(row=row_idx, column=col_res, value="REVISIÓN MANUAL PENDIENTE")
            pendientes_manual += 1

        # Guardado progresivo cada 5 casos
        if evaluados % 5 == 0:
            wb.save(AUDITED_FILE)

    wb.save(EXCEL_FILE)
    wb.save(AUDITED_FILE)

    logger.info("\n" + "=" * 75)
    logger.info("✅ EVALUACIÓN DE TRANSCRIPCIONES FINALIZADA EXITOSAMENTE:")
    logger.info(f"   • Total Casos Evaluados          : {evaluados}")
    logger.info(f"   • Con Transcripción Auditada IA  : {auditados_con_transcripcion}")
    logger.info(f"   • 🔴 Clientes que NO Aceptan     : {no_aceptados}")
    logger.info(f"   • 🟢 Clientes que Aceptan        : {aceptados}")
    logger.info(f"   • ⚪ No se Ofreció PA            : {no_ofrecidos}")
    logger.info(f"   • ⚠️ Pendientes Escucha Manual   : {pendientes_manual}")
    logger.info(f"   • Excel Principal Guardado       : {EXCEL_FILE}")
    logger.info(f"   • Excel Respaldo Guardado        : {AUDITED_FILE}")
    logger.info("=" * 75)


if __name__ == "__main__":
    main()
