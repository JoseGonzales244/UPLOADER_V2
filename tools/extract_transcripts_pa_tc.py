"""
=============================================================================
EXTRACTOR PURO (LAPTOP DE TRABAJO): GENESYS + VERINT TRANSCRIPCIONES
=============================================================================
Objetivo: 
  1. Extraer ID Llamada y Fecha Llamada desde Genesys Cloud REST API.
  2. Descargar las Transcripciones con minutaje [mm:ss] desde Verint API.
  3. Guardar las transcripciones en 'data/transcripciones_pa/' y actualizar
     'ID llamada' y 'Fecha llamada' en 'Solicitud Cumplimiento TC 2026.xlsx'.
  (NO UTILIZA LLM/GEMINI - 100% APTO PARA RED CORPORATIVA)
=============================================================================
"""
import os
import sys
import json
import time
import logging
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Optional, Dict, Any, List, Tuple
import requests
import openpyxl
import urllib3
from dotenv import load_dotenv

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

load_dotenv()

# Carpetas de Salida
OUTPUT_TRANSCRIPTS_DIR = PROJECT_ROOT / "data" / "transcripciones_pa"
OUTPUT_TRANSCRIPTS_DIR.mkdir(parents=True, exist_ok=True)

LOGS_DIR = PROJECT_ROOT / "logs"
LOGS_DIR.mkdir(parents=True, exist_ok=True)
LOG_FILE = LOGS_DIR / "extract_transcripts_pa.log"

if sys.stdout and hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

logger = logging.getLogger("ExtractTranscriptsPA")
logger.setLevel(logging.INFO)

if logger.hasHandlers():
    logger.handlers.clear()

c_handler = logging.StreamHandler(sys.stdout)
c_handler.setLevel(logging.INFO)
c_formatter = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", datefmt="%H:%M:%S")
c_handler.setFormatter(c_formatter)
logger.addHandler(c_handler)

f_handler = logging.FileHandler(LOG_FILE, encoding="utf-8", mode="a")
f_handler.setLevel(logging.DEBUG)
f_formatter = logging.Formatter("%(asctime)s [%(levelname)s] [%(filename)s:%(lineno)d] %(message)s")
f_handler.setFormatter(f_formatter)
logger.addHandler(f_handler)

from modules.verint.services.verint_api_client import VerintAPIClient
from modules.genesys.services.genesys_browser import GenesysBrowserAutomation
from modules.genesys.services.teradata_service import TeradataService
from modules.genesys.models import SolicitudAudio
from modules.genesys.config import PROFILE_DIR, GENESYS_URL

EXCEL_FILE = PROJECT_ROOT / "Solicitud Cumplimiento TC 2026.xlsx"
BACKUP_FILE = PROJECT_ROOT / "Solicitud Cumplimiento TC 2026_Auditada.xlsx"


class GenesysExtractor:
    def __init__(self, browser_bot: GenesysBrowserAutomation):
        self.browser_bot = browser_bot
        self.token: Optional[str] = None
        self.user_id_cache: Dict[str, str] = {}

    def get_token(self) -> Optional[str]:
        """Captura el Bearer Token desde Chrome CDP o perfil persistente."""
        logger.info("[GENESYS] Obteniendo Bearer Token de Genesys Cloud...")
        try:
            from playwright.sync_api import sync_playwright
            with sync_playwright() as p:
                self.browser_bot._lanzar_chrome_cdp_automatico()
                cdp_url = self.browser_bot.cdp_url.replace("localhost", "127.0.0.1")
                try:
                    browser = p.chromium.connect_over_cdp(cdp_url)
                    page = self.browser_bot._obtener_page_principal(browser)
                    if page:
                        token = self.browser_bot._extraer_bearer_token(page)
                        if token:
                            self.token = token
                            logger.info("[GENESYS] ✓ Token capturado vía CDP.")
                            return token
                except Exception as e:
                    logger.warning(f"[GENESYS] CDP no disponible: {e}. Abriendo ventana interactiva...")
                    user_data_dir = str(PROFILE_DIR)
                    PROFILE_DIR.mkdir(parents=True, exist_ok=True)
                    try:
                        context = p.chromium.launch_persistent_context(
                            user_data_dir=user_data_dir,
                            channel="chrome",
                            headless=False,
                            args=["--start-maximized"]
                        )
                    except Exception:
                        context = p.chromium.launch_persistent_context(
                            user_data_dir=user_data_dir,
                            headless=False,
                            args=["--start-maximized"]
                        )
                    page = context.pages[0] if context.pages else context.new_page()
                    page.goto(GENESYS_URL)
                    time.sleep(3)
                    token = self.browser_bot._extraer_bearer_token(page)
                    if token:
                        self.token = token
                        return token
        except Exception as e:
            logger.error(f"[GENESYS] Error en navegador: {e}")

        if not self.token:
            manual = input("\n👉 Pega tu Bearer Token de Genesys aquí (o presiona Enter para omitir): ").strip()
            if manual:
                self.token = manual.replace("Bearer ", "").strip()

        return self.token

    def search_conversation(self, reg_ev: str, telefonos: List[str], target_date: datetime) -> Tuple[Optional[str], Optional[str]]:
        if not self.token:
            return None, None

        reg_clean = str(reg_ev).strip().upper()
        if reg_clean not in self.user_id_cache:
            user_id = self.browser_bot._obtener_user_id_por_matricula(self.token, reg_clean)
            if user_id:
                self.user_id_cache[reg_clean] = user_id
        user_id = self.user_id_cache.get(reg_clean)

        d_from = (target_date - timedelta(days=2)).strftime("%Y-%m-%dT00:00:00.000Z")
        d_to = (target_date + timedelta(days=2)).strftime("%Y-%m-%dT23:59:59.000Z")
        interval = f"{d_from}/{d_to}"

        segment_filters = []
        if user_id:
            segment_filters.append({
                "type": "or",
                "predicates": [{"type": "dimension", "dimension": "userId", "operator": "matches", "value": user_id}]
            })
        if telefonos:
            dnis_preds = [{"dimension": "dnis", "value": str(tlf).strip()} for tlf in telefonos if str(tlf).strip()]
            if dnis_preds:
                segment_filters.append({"type": "or", "predicates": dnis_preds})

        if not segment_filters:
            segment_filters.append({"type": "or", "predicates": [{"dimension": "mediaType", "value": "voice"}]})

        payload = {
            "order": "desc",
            "orderBy": "conversationStart",
            "paging": {"pageSize": 25, "pageNumber": 1},
            "interval": interval,
            "segmentFilters": segment_filters
        }

        api_url = "https://api.mypurecloud.com/api/v2/analytics/conversations/details/query"
        headers = {"Authorization": f"Bearer {self.token}", "Content-Type": "application/json", "accept": "*/*"}

        try:
            resp = requests.post(api_url, headers=headers, json=payload, verify=False, timeout=15)
            if resp.status_code == 200:
                convs = resp.json().get("conversations", [])
                if convs:
                    best = convs[0]
                    return best.get("conversationId"), best.get("conversationStart")
        except Exception as e:
            logger.error(f"[GENESYS] Error en API query: {e}")

        return None, None


def main():
    if not EXCEL_FILE.exists():
        logger.critical(f"❌ No se encontró '{EXCEL_FILE}'")
        return

    logger.info("=" * 75)
    logger.info("   EXTRACTOR DE TRANSCRIPCIONES (GENESYS + VERINT)")
    logger.info(f"   Destino Transcripciones: {OUTPUT_TRANSCRIPTS_DIR}")
    logger.info("=" * 75)

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    col_map = {str(h).strip().upper(): idx + 1 for idx, h in enumerate(headers) if h}
    col_dni = col_map.get("NRO DOCUMENTO", 4)
    col_reg = col_map.get("REG_EJECUTIVO", 2)
    col_ejec = col_map.get("EJECUTIVO", 3)
    col_fec_adq = col_map.get("FECHA APROBACIN ADQ") or col_map.get("FECHA APROBACION ADQ") or 5
    col_id_llamada = col_map.get("ID LLAMADA", 8)
    col_fec_llamada = col_map.get("FECHA LLAMADA", 9)

    total_rows = ws.max_row
    total_casos = total_rows - 1

    # 1. Iniciar Genesys
    browser_bot = GenesysBrowserAutomation()
    genesys_ext = GenesysExtractor(browser_bot)
    token = genesys_ext.get_token()

    # 2. Iniciar Verint
    logger.info("[VERINT] Conectando a Verint WFO API...")
    verint_user = os.getenv("VERINT_USER")
    verint_pass = os.getenv("VERINT_PASS")
    verint_client = VerintAPIClient(username=verint_user, password=verint_pass)
    verint_ready = verint_client.login()
    if verint_ready:
        logger.info("[VERINT] ✓ Sesión en Verint activa.")
    else:
        logger.warning("[VERINT] ⚠️ No se pudo iniciar sesión en Verint API.")

    teradata_svc = TeradataService()

    transcripciones_guardadas = 0
    ids_encontrados = 0

    index_data = []

    for row_idx in range(2, total_rows + 1):
        dni_raw = ws.cell(row=row_idx, column=col_dni).value
        reg_raw = ws.cell(row=row_idx, column=col_reg).value
        ejec_raw = ws.cell(row=row_idx, column=col_ejec).value
        fec_adq_raw = ws.cell(row=row_idx, column=col_fec_adq).value
        saved_call_id = ws.cell(row=row_idx, column=col_id_llamada).value

        if not dni_raw:
            continue

        dni_8 = str(int(dni_raw) if isinstance(dni_raw, float) else dni_raw).strip().zfill(8)
        reg_clean = str(reg_raw).strip().upper() if reg_raw else ""

        fec_dt = fec_adq_raw if isinstance(fec_adq_raw, datetime) else (datetime.combine(fec_adq_raw, datetime.min.time()) if isinstance(fec_adq_raw, date) else None)
        if not fec_dt:
            try:
                fec_dt = datetime.strptime(str(fec_adq_raw)[:10], "%Y-%m-%d")
            except Exception:
                fec_dt = datetime(2026, 4, 1)

        logger.info(f"\n--- [Caso {row_idx - 1}/{total_casos}] DNI: {dni_8} | Agente: {reg_clean} ({ejec_raw}) | Fecha ADQ: {fec_dt.strftime('%Y-%m-%d')} ---")

        call_id = saved_call_id if (saved_call_id and str(saved_call_id).strip() not in ["", "7464", "None"]) else None
        fec_llamada = ws.cell(row=row_idx, column=col_fec_llamada).value

        # PASO 1: Buscar ID en Genesys si no existe
        if not call_id and token:
            dummy_sol = [SolicitudAudio(nombre_archivo=f"REQ_{dni_8}", dni=dni_8, reg_ev=reg_clean)]
            enriquecidas = teradata_svc.enriquecer_solicitudes(dummy_sol)
            telefonos = enriquecidas[0].telefonos if enriquecidas else []

            call_id, fec_llamada = genesys_ext.search_conversation(reg_clean, telefonos, fec_dt)
            if call_id:
                ws.cell(row=row_idx, column=col_id_llamada, value=str(call_id))
                if fec_llamada:
                    ws.cell(row=row_idx, column=col_fec_llamada, value=str(fec_llamada))
                ids_encontrados += 1
                logger.info(f"   ✓ ID Llamada Genesys: {call_id} ({fec_llamada})")
            else:
                logger.warning(f"   ⚠️ No encontrado en Genesys.")

        # PASO 2: Descargar Transcripción en Verint con el ID Llamada
        transcript_text = ""
        if call_id and verint_ready:
            try:
                res_data = verint_client.get_interaction_transcription_api(str(call_id).strip())
                if res_data and isinstance(res_data, dict):
                    result_obj = res_data.get("GetInteractionTranscriptionResult") or {}
                    data_obj = result_obj.get("Data") or {}
                    sequences = data_obj.get("WordsSequences") or []

                    lines = []
                    for seq in sequences:
                        if not isinstance(seq, dict):
                            continue
                        speaker = "Asesor" if seq.get("SpeakerName") == "Agent" else "Cliente"
                        start_ms = seq.get("StartTime", 0)
                        total_sec = int(start_ms) // 1000
                        mins = total_sec // 60
                        secs = total_sec % 60
                        ts_str = f"{mins:02d}:{secs:02d}"
                        words = " ".join([w.get("WordText", "") for w in seq.get("Words", []) if isinstance(w, dict) and w.get("WordText")]).strip()
                        if words:
                            lines.append(f"[{ts_str}] {speaker}: {words}")

                    if lines:
                        transcript_text = "\n".join(lines)
                        file_name = f"TRANSCRIPT_DNI_{dni_8}_{call_id}.txt"
                        out_path = OUTPUT_TRANSCRIPTS_DIR / file_name
                        with open(out_path, "w", encoding="utf-8") as f_out:
                            f_out.write(transcript_text)
                        transcripciones_guardadas += 1
                        logger.info(f"   💾 Transcripción guardada: {file_name} ({len(lines)} turnos)")
            except Exception as e_v:
                logger.error(f"   ❌ Error descargando transcripción en Verint: {e_v}")

        index_data.append({
            "fila": row_idx,
            "dni": dni_8,
            "agente": reg_clean,
            "ejecutivo": str(ejec_raw),
            "fecha_adq": fec_dt.strftime("%Y-%m-%d"),
            "id_llamada": str(call_id) if call_id else None,
            "fecha_llamada": str(fec_llamada) if fec_llamada else None,
            "archivo_transcripcion": f"TRANSCRIPT_DNI_{dni_8}_{call_id}.txt" if transcript_text else None
        })

        if (row_idx - 1) % 5 == 0:
            wb.save(EXCEL_FILE)

    wb.save(EXCEL_FILE)
    wb.save(BACKUP_FILE)

    # Guardar índice JSON
    index_file = OUTPUT_TRANSCRIPTS_DIR / "transcripciones_index.json"
    with open(index_file, "w", encoding="utf-8") as f_idx:
        json.dump(index_data, f_idx, indent=2, ensure_ascii=False)

    logger.info("\n" + "=" * 75)
    logger.info("✅ EXTRACCIÓN DE DATOS Y TRANSCRIPCIONES FINALIZADA:")
    logger.info(f"   • Total Casos Procesados          : {total_casos}")
    logger.info(f"   • IDs Identificados en Genesys    : {ids_encontrados}")
    logger.info(f"   • Transcripciones Descargadas     : {transcripciones_guardadas}")
    logger.info(f"   • Carpeta de Salida               : {OUTPUT_TRANSCRIPTS_DIR}")
    logger.info(f"   • Índice Guardado                 : {index_file}")
    logger.info(f"   • Excel Actualizado               : {EXCEL_FILE}")
    logger.info("=" * 75)


if __name__ == "__main__":
    main()
