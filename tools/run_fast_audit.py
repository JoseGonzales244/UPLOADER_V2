import os
import sys
import glob
import json
import hashlib
from pathlib import Path
import openpyxl
from dotenv import load_dotenv

PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

load_dotenv()

if sys.stdout and hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

from infrastructure.llm.gemini_client import GeminiClient

EXCEL_FILE = PROJECT_ROOT / "Solicitud Cumplimiento TC 2026.xlsx"
AUDITED_FILE = PROJECT_ROOT / "Solicitud Cumplimiento TC 2026_Auditada.xlsx"

def run():
    target_excel = AUDITED_FILE if AUDITED_FILE.exists() else EXCEL_FILE
    wb = openpyxl.load_workbook(target_excel)
    ws = wb.active

    llm = GeminiClient(default_model="gemini-3.1-flash-lite")

    cache = {}
    total_rows = ws.max_row
    print(f"Auditando {total_rows - 1} filas del Excel '{target_excel.name}'...")

    for r in range(2, total_rows + 1):
        dni = ws.cell(row=r, column=4).value
        call_id = ws.cell(row=r, column=8).value
        if not dni:
            continue
        
        dni_8 = str(int(dni) if isinstance(dni, float) else dni).strip().zfill(8)
        
        # Buscar archivo de transcripción
        matched = None
        for f in glob.glob(f"transcripciones_pa/*{dni_8}*.txt") + glob.glob(f"data/transcripciones_pa/*{dni_8}*.txt"):
            matched = f
            break
        if not matched and call_id and len(str(call_id)) > 10:
            for f in glob.glob(f"transcripciones_pa/*{call_id}*.txt") + glob.glob(f"data/transcripciones_pa/*{call_id}*.txt"):
                matched = f
                break

        if matched and os.path.exists(matched):
            with open(matched, "r", encoding="utf-8") as f_in:
                text = f_in.read().strip()
            
            if not text:
                ws.cell(row=r, column=7, value="REVISIÓN MANUAL PENDIENTE")
                continue
                
            thash = hashlib.md5(text.encode("utf-8")).hexdigest()
            if thash in cache:
                res_fmt = cache[thash]
                print(f"[{r-1}/{total_rows-1}] DNI {dni_8}: {res_fmt} (Cached)")
                ws.cell(row=r, column=7, value=res_fmt)
            else:
                prompt = f"""Eres un Auditor Senior de Cumplimiento de Televentas Bancarias de Interbank.
Determina si el asesor ofreció la AFILIACIÓN AL PAGO AUTOMÁTICO / DÉBITO AUTOMÁTICO de la Tarjeta de Crédito, y si el cliente ACEPTÓ o RECHAZÓ.

REGLAS:
1. 'NO_ACEPTA': El cliente declinó o dijo que prefiere pagar por app/banca móvil o no desea débito automático. Extrae el timestamp mm:ss.
2. 'ACEPTA': El cliente aceptó afiliarse al pago automático. Extrae el timestamp mm:ss.
3. 'NO_OFRECIDO': Nunca se mencionó en la llamada.

TRANSCRIPCIÓN:
\"\"\"
{text}
\"\"\"

Responde en JSON:
{{"estado": "NO_ACEPTA" | "ACEPTA" | "NO_OFRECIDO", "timestamp_cliente": "mm:ss" | null, "cita": "...", "explicacion": "..."}}"""
                try:
                    res_str = llm.generate_content_with_retry(prompt=prompt, model_name="gemini-3.1-flash-lite", response_json=True)
                    res_json = json.loads(res_str)
                    estado = str(res_json.get("estado")).upper()
                    ts = res_json.get("timestamp_cliente")
                    
                    if estado == "NO_ACEPTA":
                        res_fmt = f"Cliente no acepta ({ts})" if ts else "Cliente no acepta"
                    elif estado == "ACEPTA":
                        res_fmt = f"Cliente acepta ({ts})" if ts else "Cliente acepta"
                    elif estado == "NO_OFRECIDO":
                        res_fmt = "No se ofreció Pago Automático"
                    else:
                        res_fmt = "REVISIÓN MANUAL PENDIENTE"
                    
                    cache[thash] = res_fmt
                    print(f"[{r-1}/{total_rows-1}] DNI {dni_8}: {res_fmt}")
                    ws.cell(row=r, column=7, value=res_fmt)
                except Exception as e:
                    print(f"[{r-1}/{total_rows-1}] DNI {dni_8} Error LLM: {e}")
                    ws.cell(row=r, column=7, value="REVISIÓN MANUAL PENDIENTE")
        else:
            print(f"[{r-1}/{total_rows-1}] DNI {dni_8}: Sin transcripción -> REVISIÓN MANUAL PENDIENTE")
            ws.cell(row=r, column=7, value="REVISIÓN MANUAL PENDIENTE")

    wb.save(EXCEL_FILE)
    wb.save(AUDITED_FILE)
    print("\n✅ ¡AMBOS ARCHIVOS EXCEL FUERON ACTUALIZADOS Y GUARDADOS EXITOSAMENTE!")

if __name__ == "__main__":
    run()
