"""
=============================================================================
AUDITORÍA FOCALIZADA DE CUMPLIMIENTO: PAGO AUTOMÁTICO TARJETAS DE CRÉDITO
=============================================================================
Archivo Objetivo : 'Solicitud Cumplimiento TC 2026.xlsx'
Flujo de Trabajo : 
  1. Extracción de Teléfonos por DNI (zfill 8) -> Teradata / Cache Local
  2. Genesys Cloud REST API (Token CDP) -> Búsqueda conversationId y Fecha llamada
  3. Verint Speech Analytics API -> Descarga de Transcripción con marcas mm:ss
  4. Auditoría Focalizada con IA Gemini -> 'Cliente no acepta (mm:ss)'
  5. Actualización progresiva de Excel y respaldo
=============================================================================
"""
import os
import sys
import re
import json
import time
import logging
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Optional, Dict, Any, List, Tuple
import requests
import openpyxl
import urllib3
from dotenv import load_dotenv

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

PROJECT_ROOT = Path(__file__).resolve().parents[3]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

load_dotenv()

# Asegurar creación del directorio de logs
LOGS_DIR = PROJECT_ROOT / "logs"
LOGS_DIR.mkdir(parents=True, exist_ok=True)
LOG_FILE = LOGS_DIR / "audit_pago_automatico.log"

if sys.stdout and hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

logger = logging.getLogger("AuditPagoAutomatico")
logger.setLevel(logging.INFO)

if logger.hasHandlers():
    logger.handlers.clear()

c_handler = logging.StreamHandler(sys.stdout)
c_handler.setLevel(logging.INFO)
c_formatter = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", datefmt="%H:%M:%S")
c_handler.setFormatter(c_formatter)
logger.addHandler(c_handler)

f_handler = logging.FileHandler(LOG_FILE, encoding="utf-8", mode="a")
f_handler.setLevel(logging.DEBUG)
f_formatter = logging.Formatter("%(asctime)s [%(levelname)s] [%(filename)s:%(lineno)d] %(message)s")
f_handler.setFormatter(f_formatter)
logger.addHandler(f_handler)

from modules.verint.services.verint_api_client import VerintAPIClient
from infrastructure.llm.gemini_client import GeminiClient
from modules.genesys.services.genesys_browser import GenesysBrowserAutomation
from modules.genesys.services.teradata_service import TeradataService
from modules.genesys.models import SolicitudAudio

def get_default_excel_file() -> Path:
    candidates = list(PROJECT_ROOT.glob("Solicitud Cumplimiento TC*.xlsx"))
    for c in candidates:
        if "_Auditada" not in c.name:
            return c
    return PROJECT_ROOT / f"Solicitud Cumplimiento TC {datetime.now().year}.xlsx"

EXCEL_FILE = get_default_excel_file()
BACKUP_FILE = EXCEL_FILE.parent / f"{EXCEL_FILE.stem}_Auditada.xlsx"


from modules.genesys.config import PROFILE_DIR, GENESYS_URL, CDP_URL

class GenesysAPIResolver:
    """Resuelve conversationId y fecha de llamada mediante la API REST v2 de Genesys."""

    def __init__(self, browser_bot: GenesysBrowserAutomation):
        self.browser_bot = browser_bot
        self.token: Optional[str] = None
        self.user_id_cache: Dict[str, str] = {}

    def connect_and_get_token(self) -> Optional[str]:
        """Obtiene el Bearer Token activo replicando el mecanismo del módulo Genesys."""
        logger.info("[GENESYS] Inicializando conexión con Chrome / Genesys Cloud...")
        try:
            from playwright.sync_api import sync_playwright
            with sync_playwright() as p:
                browser = None
                page = None

                # 1. Auto-lanzar Chrome CDP con perfil persistente si no está activo
                if self.browser_bot._lanzar_chrome_cdp_automatico():
                    try:
                        # Reemplazar localhost por 127.0.0.1 para evitar ECONNREFUSED en IPv6
                        cdp_url = self.browser_bot.cdp_url.replace("localhost", "127.0.0.1")
                        browser = p.chromium.connect_over_cdp(cdp_url)
                        logger.info(f"[GENESYS] ✓ Conectado a Chrome vía CDP ({cdp_url})")
                        page = self.browser_bot._obtener_page_principal(browser)
                    except Exception as e:
                        logger.warning(f"[GENESYS] Error en conexión CDP: {e}")

                # 2. Fallback a persistent context si no se conectó por CDP
                if not page:
                    user_data_dir = str(PROFILE_DIR)
                    PROFILE_DIR.mkdir(parents=True, exist_ok=True)
                    logger.info(f"[GENESYS] Abriendo Chrome con perfil persistente ({PROFILE_DIR.name})...")
                    try:
                        context = p.chromium.launch_persistent_context(
                            user_data_dir=user_data_dir,
                            channel="chrome",
                            headless=False,
                            args=["--start-maximized"]
                        )
                    except Exception:
                        context = p.chromium.launch_persistent_context(
                            user_data_dir=user_data_dir,
                            headless=False,
                            args=["--start-maximized"]
                        )
                    page = context.pages[0] if context.pages else context.new_page()
                    page.goto(GENESYS_URL)
                    time.sleep(3)

                # 3. Detectar pantalla de login y esperar autenticación
                if page:
                    def _es_url_login(url_str: str) -> bool:
                        u = url_str.lower()
                        return any(d in u for d in ["microsoftonline.com", "login.live.com", "accounts.google.com", "login.windows.net"]) or "/login" in u or "login?" in u

                    if _es_url_login(page.url):
                        logger.info("🔑 [GENESYS] Sesión no iniciada o expirada.")
                        logger.info("👉 Por favor inicia sesión en la ventana de Chrome que se ha abierto (esperando hasta 5 min)...")
                        start_time = time.time()
                        while time.time() - start_time < 300:
                            if page.is_closed():
                                logger.warning("[GENESYS] Ventana cerrada por el usuario.")
                                break
                            try:
                                curr = page.url.lower()
                                if not _es_url_login(curr) and ("purecloud" in curr or "genesys" in curr or "mypurecloud" in curr):
                                    logger.info("[GENESYS] ✓ Login completado exitosamente.")
                                    break
                            except Exception:
                                pass
                            time.sleep(2)

                    token = self.browser_bot._extraer_bearer_token(page)
                    if token:
                        self.token = token
                        logger.info("[GENESYS] ✓ Bearer Token capturado exitosamente.")
                        return token
                    else:
                        logger.warning("[GENESYS] No se pudo extraer el Bearer Token de la página.")
        except Exception as e:
            logger.error(f"[GENESYS] Error en inicialización de navegador: {e}", exc_info=True)
        return None

    def resolve_user_id(self, reg_ev: str) -> Optional[str]:
        """Obtiene el GUID de usuario de Genesys para una matrícula."""
        if not self.token or not reg_ev:
            return None
        reg_clean = str(reg_ev).strip().upper()
        if reg_clean in self.user_id_cache:
            return self.user_id_cache[reg_clean]

        user_id = self.browser_bot._obtener_user_id_por_matricula(self.token, reg_clean)
        if user_id:
            self.user_id_cache[reg_clean] = user_id
        return user_id

    def search_conversation(self, user_id: Optional[str], telefonos: List[str], target_date: datetime) -> Tuple[Optional[str], Optional[str]]:
        """
        Consulta la API REST de Genesys con interval +- 2 días y retorna (conversationId, conversationStart).
        """
        if not self.token:
            return None, None

        d_from = (target_date - timedelta(days=2)).strftime("%Y-%m-%dT00:00:00.000Z")
        d_to = (target_date + timedelta(days=2)).strftime("%Y-%m-%dT23:59:59.000Z")
        interval = f"{d_from}/{d_to}"

        segment_filters = []
        if user_id:
            segment_filters.append({
                "type": "or",
                "predicates": [{"type": "dimension", "dimension": "userId", "operator": "matches", "value": user_id}]
            })

        if telefonos:
            dnis_preds = [{"dimension": "dnis", "value": str(tlf).strip()} for tlf in telefonos if str(tlf).strip()]
            if dnis_preds:
                segment_filters.append({"type": "or", "predicates": dnis_preds})

        if not segment_filters:
            segment_filters.append({
                "type": "or",
                "predicates": [{"dimension": "mediaType", "value": "voice"}]
            })

        payload = {
            "order": "desc",
            "orderBy": "conversationStart",
            "paging": {"pageSize": 50, "pageNumber": 1},
            "interval": interval,
            "segmentFilters": segment_filters
        }

        api_url = "https://api.mypurecloud.com/api/v2/analytics/conversations/details/query"
        headers = {
            "Authorization": f"Bearer {self.token}",
            "Content-Type": "application/json",
            "accept": "*/*"
        }

        try:
            resp = requests.post(api_url, headers=headers, json=payload, verify=False, timeout=15)
            if resp.status_code == 200:
                data = resp.json()
                convs = data.get("conversations", [])
                if convs:
                    # Seleccionar la conversación más larga o más relevante
                    best_conv = convs[0]
                    conv_id = best_conv.get("conversationId")
                    conv_start = best_conv.get("conversationStart")
                    logger.info(f"[GENESYS] ✓ Conversación localizada: {conv_id} ({conv_start})")
                    return conv_id, conv_start
                else:
                    logger.debug(f"[GENESYS] 0 conversaciones devueltas en interval {interval}.")
            else:
                logger.warning(f"[GENESYS] API Query retornó status {resp.status_code}: {resp.text[:200]}")
        except Exception as e:
            logger.error(f"[GENESYS] Error en query REST: {e}", exc_info=True)

        return None, None


class PagoAutomaticoAuditor:
    def __init__(self):
        self.verint_user = os.getenv("VERINT_USER")
        self.verint_pass = os.getenv("VERINT_PASS")
        self.verint_client: Optional[VerintAPIClient] = None
        self.llm_client: Optional[GeminiClient] = None

    def init_services(self) -> bool:
        """Inicializa conexiones a Verint API y cliente LLM."""
        logger.info("[INIT] Inicializando cliente LLM Gemini...")
        try:
            self.llm_client = GeminiClient(default_model="gemini-3.1-flash-lite")
            logger.info("[INIT] ✓ Cliente LLM Gemini configurado.")
        except Exception as e:
            logger.error(f"[INIT] ❌ Error inicializando Gemini: {e}")

        logger.info("[INIT] Conectando a Verint API...")
        if self.verint_user and self.verint_pass:
            try:
                self.verint_client = VerintAPIClient(username=self.verint_user, password=self.verint_pass)
                if self.verint_client.login():
                    logger.info("[INIT] ✓ Sesión en Verint API activa y autenticada.")
                    return True
                else:
                    logger.error("[INIT] ❌ Fallo de autenticación en Verint API.")
            except Exception as e:
                logger.error(f"[INIT] ❌ Excepción en Verint API: {e}", exc_info=True)
        return False

    def get_transcript_by_call_id(self, call_id: str) -> List[Tuple[str, str, str]]:
        """
        Descarga la transcripción en Verint usando el conversationId (UUID) exacto.
        """
        if not self.verint_client or not call_id:
            return []

        try:
            res_data = self.verint_client.get_interaction_transcription_api(call_id)
            if not res_data or not isinstance(res_data, dict):
                return []

            result_obj = res_data.get("GetInteractionTranscriptionResult")
            if not isinstance(result_obj, dict) or not result_obj.get("Success", False):
                return []

            data_obj = result_obj.get("Data")
            if not isinstance(data_obj, dict):
                return []

            sequences = data_obj.get("WordsSequences")
            if not isinstance(sequences, list):
                return []

            transcript_tuples = []
            for seq in sequences:
                if not isinstance(seq, dict):
                    continue
                speaker_raw = seq.get("SpeakerName", "")
                speaker = "Asesor" if speaker_raw == "Agent" else ("Cliente" if speaker_raw == "Customer" else speaker_raw)
                start_ms = seq.get("StartTime", 0)
                total_sec = int(start_ms) // 1000
                mins = total_sec // 60
                secs = total_sec % 60
                ts_str = f"{mins:02d}:{secs:02d}"

                words_list = [w.get("WordText", "") for w in seq.get("Words", []) if isinstance(w, dict) and w.get("WordText")]
                text = " ".join(words_list).strip()
                if text:
                    transcript_tuples.append((ts_str, speaker, text))

            logger.info(f"[TRANSCRIPT] ✓ Transcripción obtenida ({len(transcript_tuples)} turnos) para call_id={call_id}")
            return transcript_tuples
        except Exception as e:
            logger.error(f"[TRANSCRIPT] ❌ Error extrayendo transcripción para {call_id}: {e}", exc_info=True)
            return []

    def evaluate_pago_automatico(self, transcript_tuples: List[Tuple[str, str, str]]) -> Dict[str, Any]:
        """Audita el diálogo buscando si se ofreció y si el cliente aceptó/rechazó Pago Automático."""
        if not transcript_tuples:
            return {
                "estado": "SIN_TRANSCRIPCION",
                "timestamp": None,
                "resultado_formateado": "REVISIÓN MANUAL PENDIENTE",
                "cita": ""
            }

        full_text = "\n".join([f"[{ts}] {spk}: {txt}" for ts, spk, txt in transcript_tuples])

        prompt = f"""Eres un Auditor Senior de Cumplimiento de Televentas Bancarias de Interbank.
Tu ÚNICO objetivo es auditar la llamada para verificar si el asesor ofreció la AFILIACIÓN AL PAGO AUTOMÁTICO / DÉBITO AUTOMÁTICO de la Tarjeta de Crédito, y si el cliente ACEPTÓ o RECHAZÓ.

REGLAS DE EVALUACIÓN:
1. "NO_ACEPTA": El asesor ofreció afiliar a Pago/Débito Automático (o cargo a cuenta de ahorros), pero el cliente declinó, dijo que prefiere pagarlo por su cuenta/app, que no desea débito automático, etc.
   -> Captura el timestamp exacto (mm:ss) del momento en que el CLIENTE rechaza o dice que no.
2. "ACEPTA": El asesor ofreció Pago Automático y el cliente dio su consentimiento explícito ("Sí", "De acuerdo", "Afílieme", "Claro").
   -> Captura el timestamp exacto (mm:ss) del momento en que el CLIENTE acepta.
3. "NO_OFRECIDO": En toda la llamada NUNCA se mencionó la afiliación a Pago Automático ni Débito Automático.
   -> Timestamp: null.

TRANSCRIPCIÓN DE LA LLAMADA:
\"\"\"
{full_text}
\"\"\"

Responde estrictamente en formato JSON:
{{
  "estado": "NO_ACEPTA" | "ACEPTA" | "NO_OFRECIDO",
  "timestamp_cliente": "mm:ss" | null,
  "cita_textual_cliente": "Texto exacto donde el cliente responde o rechaza",
  "cita_textual_asesor": "Texto exacto donde el asesor ofrece el pago automático",
  "explicacion": "Breve justificación de 1 línea"
}}
"""
        if self.llm_client:
            try:
                response_str = self.llm_client.generate_content_with_retry(
                    prompt=prompt,
                    model_name="gemini-3.1-flash-lite",
                    temperature=0.0,
                    response_json=True
                )
                data = json.loads(response_str)
                estado = str(data.get("estado", "INCIERTO")).upper()
                ts = data.get("timestamp_cliente")

                if estado == "NO_ACEPTA":
                    res_fmt = f"Cliente no acepta ({ts})" if ts else "Cliente no acepta"
                elif estado == "ACEPTA":
                    res_fmt = f"Cliente acepta ({ts})" if ts else "Cliente acepta"
                elif estado == "NO_OFRECIDO":
                    res_fmt = "No se ofreció Pago Automático"
                else:
                    res_fmt = "REVISIÓN MANUAL PENDIENTE"

                logger.info(f"[LLM_EVAL] Estado: {estado} | Timestamp: {ts} | Dictamen: {res_fmt}")
                return {
                    "estado": estado,
                    "timestamp": ts,
                    "resultado_formateado": res_fmt,
                    "cita": data.get("cita_textual_cliente", ""),
                    "explicacion": data.get("explicacion", "")
                }
            except Exception as e:
                logger.error(f"[LLM_EVAL] ❌ Error en LLM: {e}", exc_info=True)

        return {
            "estado": "INCIERTO",
            "timestamp": None,
            "resultado_formateado": "REVISIÓN MANUAL PENDIENTE",
            "cita": ""
        }


def process_audit(mode: str = "full", excel_path: Optional[Path] = None):
    start_time_exec = time.time()
    target_excel = Path(excel_path) if excel_path else EXCEL_FILE
    target_backup = target_excel.parent / f"{target_excel.stem}_Auditada.xlsx"

    if not target_excel.exists():
        logger.critical(f"❌ No se encontró el archivo '{target_excel}'")
        return

    logger.info("=" * 75)
    logger.info(f"   INICIANDO AUDITORÍA FOCALIZADA: PAGO AUTOMÁTICO TC (Modo: {mode.upper()})")
    logger.info(f"   Archivo Entrada : {target_excel.name}")
    logger.info(f"   Archivo Respaldo: {target_backup.name}")
    logger.info("=" * 75)

    wb = openpyxl.load_workbook(target_excel)
    ws = wb.active

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    col_map = {str(h).strip().upper(): idx + 1 for idx, h in enumerate(headers) if h}
    col_dni = col_map.get("NRO DOCUMENTO", 4)
    col_reg = col_map.get("REG_EJECUTIVO", 2)
    col_ejec = col_map.get("EJECUTIVO", 3)
    col_fec_adq = col_map.get("FECHA APROBACIN ADQ") or col_map.get("FECHA APROBACION ADQ") or 5
    col_res = col_map.get("RESULTADO", 7)
    col_id_llamada = col_map.get("ID LLAMADA", 8)
    col_fec_llamada = col_map.get("FECHA LLAMADA", 9)

    total_rows = ws.max_row
    total_casos = total_rows - 1

    ids_encontrados_genesys = 0

    # =========================================================================
    # 📌 FASE 1: BARRIDO EN GENESYS CLOUD (Omitido en modo 'eval')
    # =========================================================================
    if mode in ["full", "extract"]:
        logger.info("\n" + "=" * 75)
        logger.info("🚀 [FASE 1] BARRIDO EN GENESYS CLOUD: OBTENIENDO IDs DE LLAMADA")
        logger.info("=" * 75)

        browser_bot = GenesysBrowserAutomation()
        genesys_resolver = GenesysAPIResolver(browser_bot)
        genesys_token = genesys_resolver.connect_and_get_token()

        teradata_svc = TeradataService()

        if genesys_token:
            for row_idx in range(2, total_rows + 1):
                dni_raw = ws.cell(row=row_idx, column=col_dni).value
                reg_raw = ws.cell(row=row_idx, column=col_reg).value
                ejec_raw = ws.cell(row=row_idx, column=col_ejec).value
                fec_adq_raw = ws.cell(row=row_idx, column=col_fec_adq).value
                saved_call_id = ws.cell(row=row_idx, column=col_id_llamada).value

                if not dni_raw:
                    continue

                dni_8 = str(int(dni_raw) if isinstance(dni_raw, float) else dni_raw).strip().zfill(8)
                reg_clean = str(reg_raw).strip().upper() if reg_raw else ""

                if saved_call_id and str(saved_call_id).strip() not in ["", "7464", "None"]:
                    ids_encontrados_genesys += 1
                    continue

                fec_dt = fec_adq_raw if isinstance(fec_adq_raw, datetime) else (datetime.combine(fec_adq_raw, datetime.min.time()) if isinstance(fec_adq_raw, date) else None)
                if not fec_dt and fec_adq_raw:
                    try:
                        fec_dt = datetime.strptime(str(fec_adq_raw)[:10], "%Y-%m-%d")
                    except Exception as parse_err:
                        logger.warning(f"No se pudo parsear fecha '{fec_adq_raw}' para DNI {dni_8}: {parse_err}")
                        fec_dt = None

                if not fec_dt:
                    logger.warning(f"⚠️ Saltando búsqueda Genesys para DNI {dni_8}: Fecha de adquisición no disponible.")
                    continue

                logger.info(f"[FASE 1 - Caso {row_idx - 1}/{total_casos}] Consultando Genesys para DNI {dni_8} | Agente: {reg_clean} ({ejec_raw})...")

                dummy_sol = [SolicitudAudio(nombre_archivo=f"REQ_{dni_8}", dni=dni_8, reg_ev=reg_clean)]
                enriquecidas = teradata_svc.enriquecer_solicitudes(dummy_sol)
                telefonos = enriquecidas[0].telefonos if enriquecidas else []

                user_id = genesys_resolver.resolve_user_id(reg_clean)
                call_id, fec_llamada = genesys_resolver.search_conversation(user_id, telefonos, fec_dt)

                if call_id:
                    ws.cell(row=row_idx, column=col_id_llamada, value=str(call_id))
                    if fec_llamada:
                        ws.cell(row=row_idx, column=col_fec_llamada, value=str(fec_llamada))
                    ids_encontrados_genesys += 1
                    logger.info(f"   ✓ Hallado CallID: {call_id} (Fecha: {fec_llamada})")
                else:
                    logger.warning(f"   ⚠️ No se halló llamada en Genesys para DNI {dni_8}.")

            wb.save(target_excel)
            logger.info(f"\n[FASE 1 COMPLETADA] {ids_encontrados_genesys}/{total_casos} IDs de llamada listos en Excel.")
        else:
            logger.warning("[FASE 1] ⚠️ No se pudo obtener sesión de Genesys. Continuando con IDs previamente guardados...")

    if mode == "extract":
        wb.save(target_excel)
        logger.info(f"✅ Modo extracción finalizado. Archivo guardado en: {target_excel}")
        return

    # =========================================================================
    # 📌 FASE 2: EXTRACCIÓN EN VERINT API Y AUDITORÍA IA
    # =========================================================================
    logger.info("\n" + "=" * 75)
    logger.info("🎯 [FASE 2] DESCARGA DE TRANSCRIPCIONES EN VERINT API Y EVALUACIÓN IA")
    logger.info("=" * 75)

    auditor = PagoAutomaticoAuditor()
    verint_ready = auditor.init_services()

    auditados_exito = 0
    pendientes_manual = 0

    for row_idx in range(2, total_rows + 1):
        dni_raw = ws.cell(row=row_idx, column=col_dni).value
        call_id = ws.cell(row=row_idx, column=col_id_llamada).value

        if not dni_raw:
            continue

        dni_8 = str(int(dni_raw) if isinstance(dni_raw, float) else dni_raw).strip().zfill(8)
        logger.info(f"\n[FASE 2 - Caso {row_idx - 1}/{total_casos}] Auditando DNI {dni_8} | ID Llamada: {call_id}...")

        transcript = []
        if call_id and str(call_id).strip() not in ["", "7464", "None"] and verint_ready:
            transcript = auditor.get_transcript_by_call_id(str(call_id).strip())

        if transcript:
            eval_res = auditor.evaluate_pago_automatico(transcript)
            resultado_texto = eval_res["resultado_formateado"]
            logger.info(f"🎯 Dictamen IA: {resultado_texto} (Cita: '{eval_res.get('cita', '')}')")

            ws.cell(row=row_idx, column=col_res, value=resultado_texto)
            auditados_exito += 1
        else:
            logger.warning(f"⚠️ Sin transcripción en Verint para DNI {dni_8} (call_id={call_id}).")
            ws.cell(row=row_idx, column=col_res, value="REVISIÓN MANUAL PENDIENTE")
            pendientes_manual += 1

        if (row_idx - 1) % 5 == 0:
            wb.save(target_excel)

    # Guardado final
    wb.save(target_excel)
    wb.save(target_backup)

    elapsed = time.time() - start_time_exec
    logger.info("\n" + "=" * 75)
    logger.info("✅ AUDITORÍA FINALIZADA:")
    logger.info(f"   • Modo de Ejecución         : {mode.upper()}")
    logger.info(f"   • Tiempo Total              : {elapsed:.1f} s")
    logger.info(f"   • Total Casos               : {total_casos}")
    logger.info(f"   • IDs Recuperados Genesys   : {ids_encontrados_genesys}")
    logger.info(f"   • Auditados con Éxito (IA)  : {auditados_exito}")
    logger.info(f"   • Pendientes Escucha Manual : {pendientes_manual}")
    logger.info(f"   • Excel Actualizado         : {target_excel}")
    logger.info(f"   • Respaldo                  : {target_backup}")
    logger.info(f"   • Log Completo              : {LOG_FILE}")
    logger.info("=" * 75)


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Auditor Unificado de Cumplimiento Pago Automático TC.")
    parser.add_argument("--mode", choices=["full", "extract", "eval", "fast"], default="full", help="Modo de ejecución: full (todo), extract (solo IDs/Verint), eval (solo IA), fast (rápido)")
    parser.add_argument("--excel", type=str, default=None, help="Ruta opcional al archivo Excel de cumplimiento")
    args = parser.parse_args()

    process_audit(mode=args.mode, excel_path=args.excel)


if __name__ == "__main__":
    main()
