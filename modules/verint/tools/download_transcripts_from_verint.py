import os
import sys
import json
import logging
from pathlib import Path
from dotenv import load_dotenv
import openpyxl

BASE_DIR = Path(__file__).resolve().parents[3]
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

from modules.verint.services.verint_api_client import VerintAPIClient

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S"
)
logger = logging.getLogger("VerintDownloader")


def main():
    load_dotenv()
    
    # Soporte para pasar archivo Excel y carpeta de salida por argumento CLI
    if len(sys.argv) > 1 and sys.argv[1].strip():
        excel_file = Path(sys.argv[1].strip())
    else:
        excel_audit = BASE_DIR / "Solicitud Cumplimiento TC 2026_Auditada.xlsx"
        excel_main = BASE_DIR / "Solicitud Cumplimiento TC 2026.xlsx"
        excel_file = excel_audit if excel_audit.exists() else excel_main
    
    if not excel_file.exists():
        logger.error(f"No se encontró el archivo Excel en: {excel_file}")
        return

    if len(sys.argv) > 2 and sys.argv[2].strip():
        output_dir = Path(sys.argv[2].strip())
    else:
        output_dir = BASE_DIR / "data" / "transcripciones_pa"
    output_dir.mkdir(parents=True, exist_ok=True)

    logger.info("=" * 75)
    logger.info(f"🎧 EXTRACTOR DIRECTO 100% VERINT WFO (DESCARGA DE TRANSCRIPCIONES)")
    logger.info(f"📂 Archivo Base : {excel_file.name}")
    logger.info(f"📁 Carpeta Salida: {output_dir}")
    logger.info("=" * 75)

    # 1. Cargar Excel
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    ws = wb.active

    header = [cell.value for cell in ws[1]]
    col_map = {}
    for idx, col_name in enumerate(header, 1):
        if col_name:
            col_map[str(col_name).strip().upper()] = idx

    col_dni = col_map.get("NRO DOCUMENTO", 4)
    col_id_llamada = col_map.get("ID LLAMADA", 8)
    col_fec_llamada = col_map.get("FECHA LLAMADA", 9)
    col_reg = col_map.get("REG_EJECUTIVO", 2)
    col_ejec = col_map.get("EJECUTIVO", 3)
    col_fec_adq = col_map.get("FECHA APROBACIN ADQ") or col_map.get("FECHA APROBACION ADQ") or 5

    total_rows = ws.max_row
    casos = []

    for row_idx in range(2, total_rows + 1):
        dni_raw = ws.cell(row=row_idx, column=col_dni).value
        call_id_raw = ws.cell(row=row_idx, column=col_id_llamada).value
        fec_llamada = ws.cell(row=row_idx, column=col_fec_llamada).value
        reg_raw = ws.cell(row=row_idx, column=col_reg).value
        ejec_raw = ws.cell(row=row_idx, column=col_ejec).value
        fec_adq_raw = ws.cell(row=row_idx, column=col_fec_adq).value

        if not dni_raw:
            continue

        dni_8 = str(int(dni_raw) if isinstance(dni_raw, float) else dni_raw).strip().zfill(8)
        call_id_str = str(call_id_raw).strip() if call_id_raw and str(call_id_raw).strip() not in ["", "7464", "None"] else None

        casos.append({
            "fila": row_idx,
            "dni": dni_8,
            "id_llamada": call_id_str,
            "fecha_llamada": str(fec_llamada) if fec_llamada else None,
            "agente": str(reg_raw).strip().upper() if reg_raw else "",
            "ejecutivo": str(ejec_raw) if ejec_raw else "",
            "fecha_adq": str(fec_adq_raw)[:10] if fec_adq_raw else None
        })

    logger.info(f"Se encontraron {len(casos)} filas en el Excel.")
    con_id = [c for c in casos if c["id_llamada"]]
    logger.info(f"Casos con ID de llamada listos para Verint: {len(con_id)}/{len(casos)}\n")

    # 2. Conexión a Verint
    user = os.getenv("VERINT_USER")
    pwd = os.getenv("VERINT_PASS")
    client = VerintAPIClient(username=user, password=pwd)
    
    if not client.login():
        logger.error("❌ No se pudo autenticar en Verint WFO. Verifica credenciales / VPN.")
        return

    logger.info("✓ Autenticación exitosa en Verint WFO.")
    logger.info("Iniciando barrido caso por caso con sesión limpia...\n")

    descargadas = 0
    sin_datos = 0

    for idx, c in enumerate(casos, 1):
        dni = c["dni"]
        call_id = c["id_llamada"]

        if not call_id:
            logger.warning(f"[{idx}/{len(casos)}] DNI: {dni} | ⚠️ Sin ID de llamada asignado. Omitiendo.")
            c["archivo_transcripcion"] = None
            continue

        logger.info(f"[{idx}/{len(casos)}] DNI: {dni} | ID Llamada: {call_id}...")

        try:
            res_data = client.get_interaction_transcription_api(call_id)
            transcript_text = VerintAPIClient.format_dialogue(res_data)
            if transcript_text:
                file_name = f"TRANSCRIPT_DNI_{dni}_{call_id}.txt"
                file_path = output_dir / file_name
                with open(file_path, "w", encoding="utf-8") as f_out:
                    f_out.write(transcript_text)
                
                c["archivo_transcripcion"] = file_name
                descargadas += 1
                first_line = transcript_text.splitlines()[0] if transcript_text.splitlines() else ""
                logger.info(f"   💾 GUARDADO: {file_name} ({len(transcript_text.splitlines())} turnos)")
                logger.info(f"   💬 Inicio: {first_line[:100]}...")
            else:
                logger.warning(f"   ⚠️ Verint no devolvió diálogo para {call_id}.")
                c["archivo_transcripcion"] = None
                sin_datos += 1
        except Exception as e:
            logger.error(f"   ❌ Excepción al procesar llamada {call_id}: {e}")
            c["archivo_transcripcion"] = None
            sin_datos += 1

    # Guardar índice final
    index_path = output_dir / "transcripciones_index.json"
    with open(index_path, "w", encoding="utf-8") as f_idx:
        json.dump(casos, f_idx, indent=2, ensure_ascii=False)

    logger.info("\n" + "=" * 75)
    logger.info("🎯 RESULTADO DE LA DESCARGA EN VERINT:")
    logger.info(f"   • Total Casos Evaluados       : {len(casos)}")
    logger.info(f"   • Transcripciones Descargadas : {descargadas}")
    logger.info(f"   • Sin Registro / No Hallados  : {sin_datos}")
    logger.info(f"   • Índice Guardado en         : {index_path}")
    logger.info("=" * 75)


if __name__ == "__main__":
    main()
