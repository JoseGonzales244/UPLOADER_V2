import os
import datetime
import openpyxl

MONTH_NAMES_UPPER = {
    1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL", 5: "MAYO", 6: "JUNIO",
    7: "JULIO", 8: "AGOSTO", 9: "SETIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
}

ANALYST_PATTERNS = {
    "CAROLINA": ["ANGULO GASTELU ANDREA CAROLINA", "ANDREA CAROLINA"],
    "JANE": ["LOPEZ ALTAMIRANO JANESY MYGHET", "JANESY", "JANE"],
    "CARMEN": ["GRADOS VILCHEZ CARMEN AMALIA", "CARMEN"],
    "KARIN": ["MAYHUASCA MATOS KARIN YOVANNA", "KARIN"]
}

def resolve_vacation_file(vacation_file):
    """Resuelve la ruta del archivo de vacaciones manejando variaciones de codificación Unicode."""
    if os.path.exists(vacation_file):
        return vacation_file
    
    parent_dir = os.path.dirname(vacation_file) or "."
    if os.path.exists(parent_dir):
        for fname in os.listdir(parent_dir):
            if fname.lower().endswith(".xlsx") and "vacaciones" in fname.lower():
                return os.path.join(parent_dir, fname)
    return vacation_file

from modules.dotacion.utils.excel import get_working_days

def get_analyst_vacations(year, month, holidays_set, vacation_file=None):
    """
    Lee 'Gestión de Vacaciones y Horarios {year}.xlsx' y retorna un diccionario con el número de días
    laborables útiles en los que cada analista tuvo vacaciones programadas (marcada con 'X') dentro del periodo
    de evaluación (excluyendo la primera semana de cierre, fines de semana y feriados nacionales).
    
    :param year: Año (ej. 2026)
    :param month: Mes numérico (ej. 8)
    :param holidays_set: conjunto (set) de objetos datetime.date con feriados del mes
    :param vacation_file: Ruta al archivo Excel de gestión de vacaciones
    :return: dict {"CAROLINA": count, "CARMEN": count, "JANE": count, "KARIN": count}
    """
    vacations_result = {"CAROLINA": 0, "CARMEN": 0, "JANE": 0, "KARIN": 0}

    if not vacation_file:
        vacation_file = f"Gestión de Vacaciones y Horarios {year}.xlsx"

    target_file = resolve_vacation_file(vacation_file)

    if not os.path.exists(target_file):
        print(f"  [WARNING] Archivo de vacaciones no encontrado. Se usará 0 días por defecto.")
        return vacations_result

    target_month_name = MONTH_NAMES_UPPER.get(month)
    if not target_month_name:
        return vacations_result

    try:
        wb = openpyxl.load_workbook(target_file, data_only=True)
        import unicodedata
        # 1. Prioridad: Hoja oficial "Programación de Fechas {year}"
        sheet_candidates = []
        for s in wb.sheetnames:
            norm = unicodedata.normalize('NFKD', s).upper()
            if str(year) in norm and 'PROGRAMA' in norm and 'FECHA' in norm:
                sheet_candidates.append(s)

        # 2. Fallbacks si la hoja tiene otro formato de nombre
        if not sheet_candidates:
            for s in wb.sheetnames:
                norm = unicodedata.normalize('NFKD', s).upper()
                if str(year) in norm and ('PROGRAMA' in norm or 'FECHA' in norm or 'TEAM' in norm):
                    sheet_candidates.append(s)
        if not sheet_candidates:
            sheet_candidates = [s for s in wb.sheetnames if "(2)" in s and str(year) in s]
        if not sheet_candidates:
            sheet_candidates = [s for s in wb.sheetnames if str(year) in s]
        
        if not sheet_candidates:
            wb.close()
            return vacations_result

        ws = wb[sheet_candidates[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if len(rows) < 4:
            return vacations_result

        row_days = rows[3]   # Fila 4

        # Mapear columnas a (NombreMes, NumeroDia) resolviendo celdas combinadas en fila 2
        row2_vals = {}
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= 2 <= rng.max_row:
                val = ws.cell(row=rng.min_row, column=rng.min_col).value
                if val:
                    for col in range(rng.min_col, rng.max_col + 1):
                        row2_vals[col] = str(val).strip().upper()

        for c in range(1, ws.max_column + 1):
            if c not in row2_vals:
                val = ws.cell(row=2, column=c).value
                if val:
                    row2_vals[c] = str(val).strip().upper()

        col_to_date = {}
        for col_idx in range(3, len(row_days)):
            col_1based = col_idx + 1
            m_name = row2_vals.get(col_1based)
            d_val = row_days[col_idx]
            if m_name and m_name in MONTH_NAMES_UPPER.values() and d_val is not None and isinstance(d_val, int):
                col_to_date[col_idx] = (m_name, d_val)

        # Ubicar filas de cada analista
        analyst_row_map = {}
        for r_idx in range(4, min(30, len(rows))):
            nom_cell = str(rows[r_idx][1] or "").upper().strip()
            for key, patterns in ANALYST_PATTERNS.items():
                if key not in analyst_row_map:
                    if any(pat in nom_cell for pat in patterns):
                        analyst_row_map[key] = r_idx

        # Calcular días útiles de evaluación para saber la fecha de inicio y fin del periodo
        working_days = get_working_days(year, month, holidays_set)
        if working_days:
            eval_start = working_days[0]
            eval_end = working_days[-1]
        else:
            eval_start = datetime.date(year, month, 1)
            eval_end = datetime.date(year, month, 28)
        
        for key, r_idx in analyst_row_map.items():
            r = rows[r_idx]
            vac_count = 0
            for col_idx, (m_name, day_num) in col_to_date.items():
                if m_name == target_month_name:
                    cell_val = str(r[col_idx] or "").strip().upper()
                    if cell_val == "X":
                        try:
                            d_obj = datetime.date(year, month, day_num)
                            if d_obj.weekday() < 5 and d_obj not in holidays_set:
                                if eval_start <= d_obj <= eval_end:
                                    vac_count += 1
                        except ValueError:
                            pass
            vacations_result[key] = vac_count

        print(f"  [OK] Vacaciones de analistas extraídas exitosamente: {vacations_result}")

    except Exception as e:
        print(f"  [WARNING] Error al procesar archivo de vacaciones: {e}")

    return vacations_result
