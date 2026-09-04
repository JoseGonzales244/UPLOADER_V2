import os
import shutil
import openpyxl
from copy import copy
from typing import Optional
from modules.dotacion.dotacion_config import DotacionConfig
from modules.dotacion.core.names import format_nom_ejecutivo_old
from modules.dotacion.utils.excel import find_headers_and_row, find_advisor_reg_col_idx

MONTH_NAMES = {
    1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
    5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
    9: "SETIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
}

SHEET_TO_SUB2 = {
    "TC": "TARJETAS",
    "SEG": "SEGUROS",
    "PP": "PRESTAMOS",
    "EC": "EXTRACASH",
    "CD": "COMPRA DE DEUDA",
    "PREHIP": "HIPOTECARIO",
    "SELECT": "SELECT",
    "CxC 1": "CONVENIOS TLV",
    "CxC": "CONVENIOS TLV",
    "RTC": "RETENCION TC",
    "RCXC": "RETENCION CONVENIOS",
    "BN_B": "BPE DESEMBOLSOS",
    "BN_C": "BPE AGENDAMIENTO"
}



def run(wb_test, cfg: Optional[DotacionConfig] = None):
    if cfg is None:
        cfg = DotacionConfig()
    # 1. Determinar nombres de archivos de ejecutivos
    prev_exec_file = cfg.PREV_EXEC_FILE
    curr_exec_file = cfg.CURR_EXEC_FILE
    out_dir = os.path.dirname(curr_exec_file) or ""

    print(f"\n--- Starting Televentas Ejecutivos synchronization for {cfg.TARGET_PERIOD} ---")
    print(f"Template path: {prev_exec_file}")
    print(f"Target file path: {curr_exec_file}")

    if not os.path.exists(prev_exec_file):
        raise FileNotFoundError(f"Template file not found: {prev_exec_file}")

    # 2. Duplicar el archivo
    print(f"[Step 1] Duplicating {prev_exec_file} to {curr_exec_file}...")
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)
    shutil.copy(prev_exec_file, curr_exec_file)

    # 3. Cargar la dotación de julio para lookups
    print("[Step 2] Indexing standard DOTACION and SELECT DOTACION for supervisor/jefe lookups...")
    dot_sheet_names = [n for n in wb_test.sheetnames if n.upper() in ["DOTACIÓN", "DOTACION"]]
    if not dot_sheet_names:
        raise ValueError("DOTACION sheet not found in workbook!")
    dot_ws = wb_test[dot_sheet_names[0]]

    headers, dot_header_row = find_headers_and_row(dot_ws, "REGISTRO COLABORADOR")
    col_reg = headers.index("REGISTRO COLABORADOR")

    standard_dot_by_reg = {}
    for r_idx in range(dot_header_row + 1, dot_ws.max_row + 1):
        reg_val = dot_ws.cell(row=r_idx, column=col_reg + 1).value
        if reg_val is not None and str(reg_val).strip() != "":
            reg = str(reg_val).strip().upper()
            row_dict = {}
            for col_idx, h in enumerate(headers):
                if h:
                    row_dict[h] = dot_ws.cell(row=r_idx, column=col_idx + 1).value
            standard_dot_by_reg[reg] = row_dict

    # Indexar hoja Dotación SELECT si existe
    select_dot_by_reg = {}
    sel_dot_sheet_names = [n for n in wb_test.sheetnames if "SELECT" in n.upper() and ("DOTACI" in n.upper() or "DOT" in n.upper())]
    if not sel_dot_sheet_names:
        # Fallback: buscar cualquier hoja que contenga SELECT excepto la hoja de producto pura
        sel_dot_sheet_names = [n for n in wb_test.sheetnames if "SELECT" in n.upper() and n.strip().upper() != "SELECT"]

    if sel_dot_sheet_names:
        sel_ws = wb_test[sel_dot_sheet_names[0]]
        sel_headers, sel_header_row = find_headers_and_row(
            sel_ws, ["REG_COLAB", "REG_PROMOTOR", "REG_EJECUTIVO", "REGISTRO"]
        )
        col_sel_reg = find_advisor_reg_col_idx(sel_headers)
        if col_sel_reg is not None:
            for r_idx in range(sel_header_row + 1, sel_ws.max_row + 1):
                reg_val = sel_ws.cell(row=r_idx, column=col_sel_reg + 1).value
                if reg_val is not None and str(reg_val).strip() != "":
                    reg = str(reg_val).strip().upper()
                    row_dict = {}
                    for col_idx, h in enumerate(sel_headers):
                        if h:
                            val = sel_ws.cell(row=r_idx, column=col_idx + 1).value
                            row_dict[h] = val
                            row_dict[str(h).strip().upper()] = val
                    select_dot_by_reg[reg] = row_dict

    # 4. Extraer asesores activos de las hojas de productos
    print("[Step 3] Extracting active advisors from product sheets...")
    active_advisors = []
    product_sheets = ["TC", "SEG", "PP", "EC", "CD", "PREHIP", "SELECT", "CxC 1", "RTC", "RCXC", "BN_B", "BN_C"]
    
    seen_regs = set()
    for s_name in product_sheets:
        if s_name not in wb_test.sheetnames:
            continue
        ws = wb_test[s_name]
        s_headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        if "REG_EV" not in s_headers:
            continue
        col_reg_idx = s_headers.index("REG_EV") + 1
        col_nombre_idx = s_headers.index("NOMBRE_EV") + 1
        col_super_idx = s_headers.index("SUPERVISOR") + 1 if "SUPERVISOR" in s_headers else None
        col_reg_super_idx = s_headers.index("REG_SUPER") + 1 if "REG_SUPER" in s_headers else None
        col_ant_idx = s_headers.index("ANTIGÜEDAD") + 1 if "ANTIGÜEDAD" in s_headers else None

        for r_idx in range(2, ws.max_row + 1):
            reg_val = ws.cell(row=r_idx, column=col_reg_idx).value
            if reg_val is not None and str(reg_val).strip() != "":
                reg_str = str(reg_val).strip().upper()
                if reg_str in seen_regs:
                    continue
                # Skip header repetitions or invalid data
                if reg_str not in ['FECHA DEL AUDIO', 'YYYYMMDD_DNI_REGISTRO_PRODUCTO_ID_PARTE', 'CAMBIO DE TCEA - 12/06/26'] and 'B12354_TC' not in reg_str:
                    seen_regs.add(reg_str)
                    
                    nombre = ws.cell(row=r_idx, column=col_nombre_idx).value
                    reg_super = ws.cell(row=r_idx, column=col_reg_super_idx).value if col_reg_super_idx else None
                    super_name = ws.cell(row=r_idx, column=col_super_idx).value if col_super_idx else None
                    ant = ws.cell(row=r_idx, column=col_ant_idx).value if col_ant_idx else "R0"
                    
                    active_advisors.append({
                        'reg': reg_str,
                        'nombre': nombre,
                        'reg_super': reg_super,
                        'super': super_name,
                        'antiguedad': ant,
                        'sheet_source': s_name
                    })
    
    print(f"  Extracted {len(active_advisors)} unique active advisors.")

    # 4.1 Auditar alertas de supervisores con códigos inconsistentes
    from modules.dotacion.core.matching import detect_supervisor_inconsistencies
    supervisor_alerts = detect_supervisor_inconsistencies(active_advisors)
    if supervisor_alerts:
        print("\n" + "=" * 70)
        print("🚨 ALERTAS DE INCONSISTENCIAS EN SUPERVISORES (INPUT EXCEL)")
        print("=" * 70)
        for alert in supervisor_alerts:
            print(f"  {alert}")
        print("=" * 70 + "\n")

    # 5. Escribir y actualizar Hoja2
    print(f"[Step 4] Updating {curr_exec_file} Hoja2...")
    wb_exec = openpyxl.load_workbook(curr_exec_file, keep_links=True, data_only=False)
    ws_exec = wb_exec['Hoja2']

    # Keep a template row to copy style
    template_row = 2
    
    # Clear Hoja2 from row 2 onwards
    for r in range(2, ws_exec.max_row + 1):
        for c in range(1, ws_exec.max_column + 1):
            ws_exec.cell(row=r, column=c).value = None

    period_val = int(cfg.TARGET_PERIOD.replace("-", ""))

    target_row = 1
    # Escribir los asesores
    for idx, adv in enumerate(active_advisors):
        target_row = 2 + idx
        
        # Copiar estilos desde la fila de plantilla original
        if target_row > template_row:
            for c in range(1, ws_exec.max_column + 1):
                src_cell = ws_exec.cell(row=template_row, column=c)
                tgt_cell = ws_exec.cell(row=target_row, column=c)
                if src_cell.has_style:
                    tgt_cell.font = copy(src_cell.font)
                    tgt_cell.border = copy(src_cell.border)
                    tgt_cell.fill = copy(src_cell.fill)
                    tgt_cell.number_format = copy(src_cell.number_format)
                    tgt_cell.alignment = copy(src_cell.alignment)

        # Mapeos de jefes y subequipos
        sub_equipo_2 = SHEET_TO_SUB2.get(adv['sheet_source'], "SELECT")
        if adv['sheet_source'] == "SELECT":
            dot_info = select_dot_by_reg.get(adv['reg']) or standard_dot_by_reg.get(adv['reg'], {})
        else:
            dot_info = standard_dot_by_reg.get(adv['reg'], {})

        # Buscar nombre de jefe con fallbacks flexibles (en Select la columna suele llamarse NOM_JEFE)
        jefe_name = (
            dot_info.get("NOM_JEFE") or
            dot_info.get("JEFE") or
            dot_info.get("JEFE DIRECTO") or
            dot_info.get("JEFE DE VENTAS") or
            dot_info.get("NOMBRE_JEFE")
        )
        subgerente_name = (
            dot_info.get("SUBGERENTE") or
            dot_info.get("NOM_SUBGERENTE") or
            dot_info.get("SUB_GERENTE")
        )

        # Si el supervisor o su registro estaban vacíos en la hoja de producto, completar desde dotación
        super_name = adv['super'] or dot_info.get("SUPERVISOR") or dot_info.get("NOM_SUP") or dot_info.get("SUPERVISOR / JEFE")
        reg_super = adv['reg_super'] or dot_info.get("REG_SUPER") or dot_info.get("REG_SUP") or dot_info.get("REGISTRO_SUPER") or dot_info.get("REG SUPERVISOR JEFE")

        # Si dot_info ya tiene REG_JEFE explícito (como en Dotación Select), usarlo directamente si el VLOOKUP fallara
        reg_jefe_direct = dot_info.get("REG_JEFE") or dot_info.get("REGISTRO_JEFE")
        formula_reg_jefe = reg_jefe_direct if reg_jefe_direct else f'=VLOOKUP(Tabla15[[#This Row],[NOM_JEFE]],JEFE[],2,FALSE)'

        # Formulas exactas de VLOOKUP y CONCAT
        formula_equipo = f'=VLOOKUP(Tabla15[[#This Row],[SUB_EQUIPO_2]],EQUIPOS[],3,FALSE)'
        formula_sub_equipo = f'=VLOOKUP(Tabla15[[#This Row],[SUB_EQUIPO_2]],EQUIPOS[],2,FALSE)'
        formula_codigo = f'=_xlfn.CONCAT(Tabla15[[#This Row],[PERIODO]],"_",Tabla15[[#This Row],[REG_EJECUTIVO]])'

        # Asignar valores
        ws_exec.cell(row=target_row, column=1, value=period_val)
        ws_exec.cell(row=target_row, column=2, value=adv['reg'])
        ws_exec.cell(row=target_row, column=3, value=adv['nombre'])
        ws_exec.cell(row=target_row, column=4, value=reg_super)
        ws_exec.cell(row=target_row, column=5, value=super_name)
        ws_exec.cell(row=target_row, column=6, value=formula_reg_jefe)
        ws_exec.cell(row=target_row, column=7, value=jefe_name)
        ws_exec.cell(row=target_row, column=8, value=formula_equipo)
        ws_exec.cell(row=target_row, column=9, value=formula_sub_equipo)
        ws_exec.cell(row=target_row, column=10, value=sub_equipo_2)
        ws_exec.cell(row=target_row, column=11, value=subgerente_name)
        ws_exec.cell(row=target_row, column=12, value=adv['antiguedad'])
        ws_exec.cell(row=target_row, column=13, value=format_nom_ejecutivo_old(adv['nombre']))
        ws_exec.cell(row=target_row, column=14, value=formula_codigo)

    # Delete excess rows if new roster is smaller than previous
    if ws_exec.max_row > target_row and target_row > 0:
        print(f"  Deleting {ws_exec.max_row - target_row} excess rows...")
        ws_exec.delete_rows(target_row + 1, ws_exec.max_row - target_row)

    # 6. Actualizar las referencias de la tabla de Excel (Tabla15)
    total_rows = 1 + len(active_advisors)
    print(f"[Step 5] Resizing table references to A1:N{total_rows}...")
    tables = getattr(ws_exec, 'tables', None) or getattr(ws_exec, '_tables', {})
    tbl_items = tables.values() if isinstance(tables, dict) else (tables if hasattr(tables, '__iter__') else [])
    for tbl in tbl_items:
        tbl_name = getattr(tbl, 'name', None) or getattr(tbl, 'displayName', None)
        if tbl_name == 'Tabla15':
            tbl.ref = f"A1:N{total_rows}"
            if getattr(tbl, 'autoFilter', None):
                tbl.autoFilter.ref = f"A1:N{total_rows}"

    try:
        wb_exec.save(curr_exec_file)
        print(f"Workbook successfully saved to: {curr_exec_file}")
    except PermissionError:
        print(f"[ERROR] No se pudo guardar el archivo {curr_exec_file}. Verifique que no esté abierto en Excel.")
        raise
    finally:
        wb_exec.close()

    print("Televentas Ejecutivos synchronization complete!")
