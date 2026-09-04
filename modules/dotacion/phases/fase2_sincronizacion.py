import os
import datetime
import openpyxl
from typing import Optional
from modules.dotacion.dotacion_config import DotacionConfig
from modules.dotacion.core.absence import check_absence_status
from modules.dotacion.core.matching import mismo_supervisor
from modules.dotacion.utils.excel import find_headers_and_row, copy_row_style, get_working_days, find_advisor_reg_col_idx

def get_new_seniority(prev_val):
    if not prev_val:
        return "R0"
    v = str(prev_val).strip().upper()
    if v == "R0":
        return "R1"
    elif v == "R1":
        return "R2"
    elif v in ["R2", "R3", "R4"]:
        return "R3"
    return v

def run(wb, cfg: Optional[DotacionConfig] = None):
    if cfg is None:
        cfg = DotacionConfig()
    TARGET_PERIOD = cfg.TARGET_PERIOD
    print(f"\n--- Starting Fase 2 Roster Synchronization for: {TARGET_PERIOD} ---")
    
    # 1. Index DOTACION sheet
    print("[Step 1] Indexing DOTACION sheets...")
    dot_sheet_names = [n for n in wb.sheetnames if n.upper() in ["DOTACIÓN", "DOTACION"]]
    if not dot_sheet_names:
        raise ValueError("DOTACION sheet not found in workbook!")
    dot_ws = wb[dot_sheet_names[0]]
    
    # Build dictionary for standard dotacion
    standard_dot_by_reg = {}
    headers, dot_header_row = find_headers_and_row(dot_ws, "REGISTRO COLABORADOR")
    col_reg = headers.index("REGISTRO COLABORADOR")
    
    # Buscar índice de columna para COL_AT si existe
    col_at_idx = None
    for idx_h, h_val in enumerate(headers):
        if h_val and "COL_AT" in str(h_val).upper():
            col_at_idx = idx_h + 1
            break

    for r_idx in range(dot_header_row + 1, dot_ws.max_row + 1):
        reg_val = dot_ws.cell(row=r_idx, column=col_reg + 1).value
        if reg_val is not None and str(reg_val).strip() != "":
            reg = str(reg_val).strip().upper()
            row_dict = {}
            for col_idx, h in enumerate(headers):
                if h:
                    row_dict[h] = dot_ws.cell(row=r_idx, column=col_idx + 1).value
            row_dict['COL_AT'] = dot_ws.cell(row=r_idx, column=col_at_idx).value if col_at_idx else dot_ws.cell(row=r_idx, column=46).value
            standard_dot_by_reg[reg] = row_dict
            
    # Index SELECT DOTACION sheet
    sel_dot_sheet_names = [n for n in wb.sheetnames if "SELECT" in n.upper() and "DOTACI" in n.upper()]
    if not sel_dot_sheet_names:
        raise ValueError("Dotacion SELECT sheet not found in workbook!")
    sel_ws = wb[sel_dot_sheet_names[0]]
    
    select_dot_by_reg = {}
    sel_headers, sel_header_row = find_headers_and_row(
        sel_ws, ["REG_COLAB", "REG_PROMOTOR", "REG_EJECUTIVO", "REGISTRO"]
    )
    col_sel_reg = find_advisor_reg_col_idx(sel_headers)
    if col_sel_reg is None:
        raise ValueError(f"Could not find advisor registration column in sheet '{sel_ws.title}'")
    
    col_sel_colab = None
    for idx_h, h_name in enumerate(sel_headers):
        if h_name and ("COLABORADOR" in str(h_name).upper() or "NOMBRE" in str(h_name).upper()):
            col_sel_colab = idx_h
            break

    for r_idx in range(sel_header_row + 1, sel_ws.max_row + 1):
        reg_val = sel_ws.cell(row=r_idx, column=col_sel_reg + 1).value
        colab_val = sel_ws.cell(row=r_idx, column=col_sel_colab + 1).value if col_sel_colab is not None else None
        if reg_val is not None and str(reg_val).strip() != "":
            reg = str(reg_val).strip().upper()
        elif colab_val and str(colab_val).strip() != "":
            reg = f"NOREG_{r_idx}"
        else:
            continue

        row_dict = {}
        for col_idx, h in enumerate(sel_headers):
            if h:
                row_dict[h] = sel_ws.cell(row=r_idx, column=col_idx + 1).value
        select_dot_by_reg[reg] = row_dict

    # Cargar el archivo del mes anterior para referencias de antigüedad
    prev_seniorities = {}
    prev_workbook_path = cfg.INPUT_WORKBOOK
    if os.path.exists(prev_workbook_path):
        print(f"Loading previous month's workbook for seniority reference: {prev_workbook_path}")
        try:
            prev_wb = openpyxl.load_workbook(prev_workbook_path, data_only=True)
            sheet_names_to_check = ["TC", "SEG", "PP", "EC", "CD", "PREHIP", "SELECT", "CxC 1", "RTC", "RCXC", "BN_B", "BN_C"]
            for sheet_name in sheet_names_to_check:
                if sheet_name in prev_wb.sheetnames:
                    p_sheet = prev_wb[sheet_name]
                    p_headers = [str(p_sheet.cell(row=1, column=c).value).strip() if p_sheet.cell(row=1, column=c).value is not None else '' for c in range(1, p_sheet.max_column+1)]
                    if "REG_EV" in p_headers:
                        r_col = p_headers.index("REG_EV") + 1
                        ant_col = p_headers.index("ANTIGÜEDAD") + 1 if "ANTIGÜEDAD" in p_headers else (p_headers.index("ANTIGUEDAD") + 1 if "ANTIGUEDAD" in p_headers else -1)
                        if ant_col != -1:
                            for r_idx in range(2, p_sheet.max_row+1):
                                reg_val = p_sheet.cell(row=r_idx, column=r_col).value
                                if reg_val is not None:
                                    reg_str = str(reg_val).strip().upper()
                                    ant_val = p_sheet.cell(row=r_idx, column=ant_col).value
                                    if ant_val is not None:
                                        prev_seniorities[(sheet_name, reg_str)] = str(ant_val).strip().upper()
            prev_wb.close()
            print(f"  Loaded {len(prev_seniorities)} seniority references from previous month.")
        except Exception as e:
            print(f"  Warning: Could not read previous workbook seniorities: {e}")
    else:
        print(f"  Warning: Previous month's workbook not found at: {prev_workbook_path}")

    # Calculate business days count from FERIADOS sheet if present
    holidays = set()
    if "FERIADOS" in wb.sheetnames:
        feriados_sheet = wb["FERIADOS"]
        for row in feriados_sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:
                if isinstance(row[0], datetime.datetime):
                    holidays.add(row[0].date())
                elif isinstance(row[0], datetime.date):
                    holidays.add(row[0])
    year, month = map(int, TARGET_PERIOD.split("-"))
    working_days_list = get_working_days(year, month, holidays)
    num_business_days = len(working_days_list)
    print(f"  Calculated working days: {num_business_days}")

    # 2. Setup product filters
    cxc_subs = ['CONVENIOS TLV', 'CONVENIOS TLV 2X', 'CONVENIOS TLV - ASISTIDO', 'CONVENIOS TLV SNIPER']
    rcxc_subs = ['RETENCION CONVENIOS', 'CONVENIOS TLV - PILOTO']
    
    SHEET_FILTERS = {
        "TC": {
            "dot_source": "standard",
            "filter": lambda r: r.get('EQUIPO') == 'TELEVENTAS TC' or r.get('SUB EQUIPO') == 'TARJETAS',
            "product_val": "TARJETA"
        },
        "SEG": {
            "dot_source": "standard",
            "filter": lambda r: r.get('EQUIPO') == 'POSTVENTA' and r.get('SUB EQUIPO') == 'SEGUROS',
            "product_val": "SEGURO"
        },
        "PP": {
            "dot_source": "standard",
            "filter": lambda r: r.get('EQUIPO') == 'TLV PRODUCTOS' and r.get('SUB EQUIPO') in ['PRESTAMOS', 'PRESTAMO COMPRA DE DEUDA'],
            "product_val": "PRESTAMO"
        },
        "EC": {
            "dot_source": "standard",
            "filter": lambda r: r.get('EQUIPO') == 'TLV PRODUCTOS' and r.get('SUB EQUIPO') == 'EXTRACASH',
            "product_val": "EXTRACASH"
        },
        "CD": {
            "dot_source": "standard",
            "filter": lambda r: r.get('EQUIPO') == 'TLV PRODUCTOS' and r.get('SUB EQUIPO') == 'COMPRA DE DEUDA',
            "product_val": "COMPRA DE DEUDA"
        },
        "PREHIP": {
            "dot_source": "standard",
            "filter": lambda r: r.get('EQUIPO') == 'TLV PRODUCTOS' and r.get('SUB EQUIPO') == 'HIPOTECARIO',
            "product_val": "HIPOTECARIO"
        },
        "SELECT": {
            "dot_source": "select",
            "filter": lambda r: True,
            "product_val": "SELECT"
        },
        "CxC 1": {
            "dot_source": "standard",
            "filter": lambda r: r.get('EQUIPO') == 'TLV PRODUCTOS' and r.get('SUB EQUIPO') in cxc_subs,
            "product_val": "CONVENIOS TLV",
            "use_sub_equipo_for_product": True
        },
        "RTC": {
            "dot_source": "standard",
            "filter": lambda r: r.get('EQUIPO') == 'TLV PRODUCTOS' and r.get('SUB EQUIPO') in ['RETENCION TC', 'RETENCION EXCLUSIVO'],
            "product_val": "RETENCION TC"
        },
        "RCXC": {
            "dot_source": "standard",
            "filter": lambda r: r.get('EQUIPO') == 'TLV PRODUCTOS' and r.get('SUB EQUIPO') in rcxc_subs,
            "product_val": "RETENCION CONVENIOS",
            "use_sub_equipo_for_product": True
        },
        "BN_B": {
            "dot_source": "standard",
            "filter": lambda r: r.get('EQUIPO') == 'TLV PRODUCTOS' and r.get('SUB EQUIPO') == 'BPE DESEMBOLSOS',
            "product_val": "BNB"
        },
        "BN_C": {
            "dot_source": "standard",
            "filter": lambda r: r.get('EQUIPO') == 'TLV PRODUCTOS' and r.get('SUB EQUIPO') == 'BPE AGENDAMIENTO',
            "product_val": "BN_C"
        }
    }

    # Pre-load registrations in CxC 1 and CxC sheets to avoid cross-sheet duplication
    cxc1_current_regs = set()
    cxc_current_regs = set()
    for cxc_name, cxc_set in [("CxC 1", cxc1_current_regs), ("CxC", cxc_current_regs)]:
        if cxc_name in wb.sheetnames:
            cxc_ws = wb[cxc_name]
            cxc_h = [cell.value for cell in next(cxc_ws.iter_rows(max_row=1))]
            if "REG_EV" in cxc_h:
                cxc_col_idx = cxc_h.index("REG_EV") + 1
                consec = 0
                for r_idx in range(2, cxc_ws.max_row + 1):
                    val = cxc_ws.cell(row=r_idx, column=cxc_col_idx).value
                    if val is not None and str(val).strip() != "":
                        consec = 0
                        reg_str = str(val).strip().upper()
                        if reg_str not in ['FECHA DEL AUDIO', 'YYYYMMDD_DNI_REGISTRO_PRODUCTO_ID_PARTE', 'CAMBIO DE TCEA - 12/06/26'] and 'B12354_TC' not in reg_str:
                            cxc_set.add(reg_str)
                    else:
                        consec += 1
                        if consec > 20:
                            break

    # Lists to keep track of updates for logging
    seniority_changes = []
    status_changes = []
    supervisor_changes = []
    added_advisors = []

    # 3. Process sheets
    print("\n[Step 2] Synchronizing rosters on product sheets...")
    for s_name, config_filter in SHEET_FILTERS.items():
        if s_name not in wb.sheetnames:
            continue
            
        print(f"  Synchronizing sheet: {s_name}")
        sheet = wb[s_name]
        
        # Read headers
        sheet_headers = [sheet.cell(row=1, column=c).value for c in range(1, sheet.max_column + 1)]
        if "REG_EV" not in sheet_headers:
            continue
            
        col_reg_idx = sheet_headers.index("REG_EV") + 1
        col_nombre_idx = sheet_headers.index("NOMBRE_EV") + 1
        col_super_idx = sheet_headers.index("SUPERVISOR") + 1 if "SUPERVISOR" in sheet_headers else None
        col_reg_super_idx = sheet_headers.index("REG_SUPER") + 1 if "REG_SUPER" in sheet_headers else None
        col_ant_idx = sheet_headers.index("ANTIGÜEDAD") + 1 if "ANTIGÜEDAD" in sheet_headers else None
        col_prod_idx = sheet_headers.index("PRODUCTO") + 1 if "PRODUCTO" in sheet_headers else None
        col_jefe_idx = sheet_headers.index("JEFE") + 1 if "JEFE" in sheet_headers else None
        
        eval_idx = None
        for idx, h in enumerate(sheet_headers):
            if h == "EVAL 1 (ID AUDIO FECHA)":
                eval_idx = idx + 1
                break
                
        eec_cols = [idx + 1 for idx, h in enumerate(sheet_headers) if h and str(h).strip().startswith("EEC")]
        
        # Load current roster advisors
        current_advisors = {}
        consec_empty = 0
        for r_idx in range(2, sheet.max_row + 1):
            val = sheet.cell(row=r_idx, column=col_reg_idx).value
            if val is not None and str(val).strip() != "":
                consec_empty = 0
                reg_str = str(val).strip().upper()
                if reg_str not in ['FECHA DEL AUDIO', 'YYYYMMDD_DNI_REGISTRO_PRODUCTO_ID_PARTE', 'CAMBIO DE TCEA - 12/06/26'] and 'B12354_TC' not in reg_str:
                    current_advisors[reg_str] = r_idx
            else:
                consec_empty += 1
                if consec_empty > 20:
                    break
                    
        # Filter new active July roster
        if config_filter["dot_source"] == "standard":
            source_dot = standard_dot_by_reg
            filtered_roster = {}
            for reg, r_dict in source_dot.items():
                if config_filter["filter"](r_dict):
                    puesto = str(r_dict.get('PUESTO', '')).upper()
                    col_at_val = str(r_dict.get('COL_AT', '')).upper()
                    estado = str(r_dict.get('ESTADO', '')).upper()
                    sub_eq = str(r_dict.get('SUB EQUIPO') or r_dict.get('SUB_EQUIPO') or '').strip().upper()
                    obs_val = str(r_dict.get('OBSERVACIONES') or '').upper()
                    if (estado in ['ACTIVO', 'LACTANCIA'] and 
                        sub_eq != 'CONVENIOS TLV - FUVEX' and
                        'SUPERVISOR' not in puesto and 
                        'JEFE' not in puesto and 
                        'GERENTE' not in puesto and 
                        'ASISTENTE BO' not in col_at_val and
                        'ASISTENTE' not in puesto and 
                        'BO' not in puesto and
                        'BACKOFFICE' not in obs_val):
                        filtered_roster[reg] = r_dict
        else:
            source_dot = select_dot_by_reg
            filtered_roster = {}
            for reg, r_dict in source_dot.items():
                puesto = str(r_dict.get('PUESTO', '')).upper()
                estado = str(r_dict.get('ESTADO', '')).upper()
                if estado != 'INACTIVO' and 'SUPERVISOR' not in puesto and 'JEFE' not in puesto:
                    filtered_roster[reg] = r_dict
        
        new_rows = []
        # 1. Collect active existing advisors
        for reg, row_num in current_advisors.items():
            if reg in filtered_roster:
                r_dict = filtered_roster[reg]
                colab_name = (
                    r_dict.get('COLABORADOR') or 
                    r_dict.get('NOMBRE_COLABORADOR') or 
                    r_dict.get('APELLIDOS Y NOMBRES') or 
                    r_dict.get('NOMBRE')
                )
                
                if config_filter["dot_source"] == "standard":
                    reg_super = r_dict.get('REG SUPERVISOR JEFE')
                    super_name = r_dict.get('SUPERVISOR / JEFE')
                    jefe_name = None
                else:
                    reg_super = r_dict.get('REGISTRO_SUPER') or r_dict.get('REG_SUP')
                    super_name = r_dict.get('SUPERVISOR') or r_dict.get('NOM_SUP')
                    jefe_name = r_dict.get('JEFE') or r_dict.get('NOM_JEFE')
                
                old_ant = sheet.cell(row=row_num, column=col_ant_idx).value if col_ant_idx else None
                old_ant_prev = prev_seniorities.get((s_name, reg))
                if old_ant_prev:
                    new_ant = get_new_seniority(old_ant_prev)
                else:
                    new_ant = str(r_dict.get('ANTIGÜEDAD') or r_dict.get('ANTIGUEDAD') or r_dict.get('ANTIGUEDAD CANAL') or 'R0').strip().upper()
                old_super = sheet.cell(row=row_num, column=col_super_idx).value if col_super_idx else None
                
                if col_ant_idx:
                    if old_ant and str(old_ant).strip().upper() != str(new_ant).strip().upper():
                        seniority_changes.append((s_name, reg, colab_name, old_ant, new_ant))
                
                if col_super_idx and super_name:
                    if old_super and not mismo_supervisor(old_super, super_name):
                        supervisor_changes.append((s_name, reg, colab_name, old_super, super_name))
                
                # Check for vacation or full month absences
                absence = None
                if config_filter["dot_source"] == "standard":
                    absence = check_absence_status(r_dict, num_business_days, holidays, year, month)
                    if absence:
                        status_changes.append((s_name, reg, colab_name, "ACTIVO", absence))
                

                
                new_rows.append({
                    'reg': reg,
                    'nombre': colab_name,
                    'antiguedad': new_ant,
                    'reg_super': reg_super,
                    'super': super_name,
                    'jefe': jefe_name,
                    'absence': absence,
                    'dias_vac': r_dict.get('DIAS VACACIONES TOTALES', 0) or 0,
                    'sub_equipo': r_dict.get('SUB EQUIPO') or r_dict.get('SUB_EQUIPO')
                })
            else:
                # Advisor is not in active roster (Cese or Traslado) -> log deletion
                is_transfer = False
                if reg in standard_dot_by_reg:
                    full_r = standard_dot_by_reg[reg]
                    if full_r.get('ESTADO') == 'ACTIVO':
                        is_transfer = True
                status_str = 'TRASLADO' if is_transfer else 'CESE'
                colab_name = sheet.cell(row=row_num, column=col_nombre_idx).value
                status_changes.append((s_name, reg, colab_name, "ACTIVO", f"{status_str} (ELIMINADO)"))
                
        # 2. Collect new advisors (Altas)
        for reg, r_dict in filtered_roster.items():
            if reg not in current_advisors:
                # Avoid cross-sheet duplication for CxC sheets
                if s_name == "CxC" and reg in cxc1_current_regs:
                    continue
                if s_name == "CxC 1" and reg in cxc_current_regs:
                    continue
                    
                colab_name = (
                    r_dict.get('COLABORADOR') or 
                    r_dict.get('NOMBRE_COLABORADOR') or 
                    r_dict.get('APELLIDOS Y NOMBRES') or 
                    r_dict.get('NOMBRE')
                )
                new_sup = r_dict.get('SUPERVISOR / JEFE') if config_filter["dot_source"] == "standard" else (r_dict.get('SUPERVISOR') or r_dict.get('NOM_SUP'))
                added_advisors.append((s_name, reg, colab_name, new_sup))
                
                if config_filter["dot_source"] == "standard":
                    reg_super = r_dict.get('REG SUPERVISOR JEFE')
                    super_name = r_dict.get('SUPERVISOR / JEFE')
                    jefe_name = None
                else:
                    reg_super = r_dict.get('REGISTRO_SUPER') or r_dict.get('REG_SUP')
                    super_name = r_dict.get('SUPERVISOR') or r_dict.get('NOM_SUP')
                    jefe_name = r_dict.get('JEFE') or r_dict.get('NOM_JEFE')
                
                old_ant_prev = prev_seniorities.get((s_name, reg))
                if old_ant_prev:
                    new_ant = get_new_seniority(old_ant_prev)
                else:
                    new_ant = str(r_dict.get('ANTIGÜEDAD') or r_dict.get('ANTIGUEDAD') or r_dict.get('ANTIGUEDAD CANAL') or 'R0').strip().upper()
                
                # Check absence for new advisors too
                absence = None
                if config_filter["dot_source"] == "standard":
                    absence = check_absence_status(r_dict, num_business_days, holidays, year, month)
                    if absence:
                        status_changes.append((s_name, reg, colab_name, "ACTIVO", absence))
                


                new_rows.append({
                    'reg': reg,
                    'nombre': colab_name,
                    'antiguedad': new_ant,
                    'reg_super': reg_super,
                    'super': super_name,
                    'jefe': jefe_name,
                    'absence': absence,
                    'dias_vac': r_dict.get('DIAS VACACIONES TOTALES', 0) or 0,
                    'sub_equipo': r_dict.get('SUB EQUIPO') or r_dict.get('SUB_EQUIPO')
                })
        
        # Helper to get product value for sorting
        def get_row_product(r_data, cfg):
            if cfg.get("use_sub_equipo_for_product") and r_data.get("sub_equipo"):
                return str(r_data["sub_equipo"] or '')
            return str(cfg["product_val"] or '')

        # Sort new_rows by PRODUCTO and SUPERVISOR if available (case-insensitive)
        def sort_key(r):
            prod_val = get_row_product(r, config_filter) if col_prod_idx else ''
            super_val = str(r.get('super') or '') if col_super_idx else ''
            name_val = str(r.get('nombre') or '')
            return (prod_val.upper(), super_val.upper(), name_val.upper())
            
        new_rows.sort(key=sort_key)

        # 3. Clear non-formula cells from row 2 down to max_clear_row
        max_clear_row = max(2 + len(new_rows), max(current_advisors.values()) if current_advisors else 2)
        for r_idx in range(2, max_clear_row + 1):
            for c_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=r_idx, column=c_idx)
                if cell.value is not None and not (isinstance(cell.value, str) and cell.value.startswith('=')):
                    cell.value = None
                    
        # 4. Write new_rows sequentially starting from row 2
        template_row_num = 2
        for idx, r_data in enumerate(new_rows):
            target_row = 2 + idx
            if target_row > 2:
                copy_row_style(sheet, template_row_num, target_row)
                
            reg_val = r_data['reg']
            if reg_val and str(reg_val).startswith('NOREG_'):
                reg_val = None
            sheet.cell(row=target_row, column=col_reg_idx).value = reg_val
            sheet.cell(row=target_row, column=col_nombre_idx).value = r_data['nombre']
            if col_ant_idx:
                ant_cell = sheet.cell(row=target_row, column=col_ant_idx)
                ant_val = str(r_data['antiguedad'] or '').strip().upper()
                ant_cell.value = ant_val
                if ant_val in ['R0', 'R1', 'R2']:
                    from openpyxl.styles import PatternFill
                    ant_cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                else:
                    from openpyxl.styles import PatternFill
                    ant_cell.fill = PatternFill(fill_type=None)
            if col_reg_super_idx and r_data['reg_super']:
                sheet.cell(row=target_row, column=col_reg_super_idx).value = r_data['reg_super']
            if col_super_idx and r_data['super']:
                sheet.cell(row=target_row, column=col_super_idx).value = r_data['super']
            if col_jefe_idx and r_data['jefe']:
                sheet.cell(row=target_row, column=col_jefe_idx).value = r_data['jefe']
            if col_prod_idx:
                if config_filter.get("use_sub_equipo_for_product") and r_data.get("sub_equipo"):
                    sheet.cell(row=target_row, column=col_prod_idx).value = r_data["sub_equipo"]
                else:
                    sheet.cell(row=target_row, column=col_prod_idx).value = config_filter["product_val"]
                
            if r_data.get('absence'):
                abs_val = str(r_data['absence']).strip()
                if abs_val == 'VACACIONES':
                    # In 2-evaluation sheets, partial vacations (< 95%) write VACACIONES to EEC 1 so EEC 2 is evaluable.
                    # In 1-evaluation sheets, partial vacations (< 95%) MUST BE EVALUATED, so do NOT write VACACIONES to EEC 1!
                    if len(eec_cols) == 2:
                        eec_map_sync = {}
                        for c_i, h in enumerate(sheet_headers, 1):
                            if h and str(h).strip().startswith("EEC"):
                                h_clean = str(h).strip().replace(" ", "").upper()
                                eec_map_sync[h_clean] = c_i
                        col_eec1 = eec_map_sync.get("EEC1") or eec_map_sync.get("EEC 1") or eec_map_sync.get("EEC")
                        if col_eec1:
                            cell_vac = sheet.cell(row=target_row, column=col_eec1)
                            cell_vac.value = 'VACACIONES'
                            from openpyxl.styles import Font
                            cell_vac.font = Font(bold=True, color="FF0000")
                else:
                    # VACACIONES TODO EL MES (>= 95% absent) or LICENCIA -> write to all EEC columns
                    for col_c in eec_cols:
                        cell = sheet.cell(row=target_row, column=col_c)
                        cell.value = abs_val
                        if abs_val.startswith('VACACIONES'):
                            from openpyxl.styles import Font
                            cell.font = Font(bold=True, color="FF0000")
                if eval_idx:
                    cell_ev = sheet.cell(row=target_row, column=eval_idx)
                    cell_ev.value = abs_val
                    if abs_val.startswith('VACACIONES'):
                        from openpyxl.styles import Font
                        cell_ev.font = Font(bold=True, color="FF0000")
            elif len(eec_cols) >= 2:
                ant_val = str(r_data.get('antiguedad') or '').strip().upper()
                try:
                    d_vac = int(r_data.get('dias_vac', 0) or 0)
                except (ValueError, TypeError):
                    d_vac = 0
                d_trabajo = num_business_days - d_vac
                
                if d_vac > 0 and d_trabajo <= 5:
                    cell_eec2 = sheet.cell(row=target_row, column=eec_cols[1])
                    cell_eec2.value = 'VACACIONES'
                    from openpyxl.styles import Font
                    cell_eec2.font = Font(
                        name=cell_eec2.font.name if cell_eec2.font else None,
                        size=cell_eec2.font.size if cell_eec2.font else None,
                        bold=True,
                        italic=cell_eec2.font.italic if cell_eec2.font else False,
                        color="FF0000"
                    )
                elif ant_val == 'R0' and s_name != 'BN_B':
                    eec_map_sync = {}
                    for c_i, h in enumerate(sheet_headers, 1):
                        if h and str(h).strip().startswith("EEC"):
                            h_clean = str(h).strip().replace(" ", "").upper()
                            eec_map_sync[h_clean] = c_i
                    col_eec1 = eec_map_sync.get("EEC1") or eec_map_sync.get("EEC 1") or eec_map_sync.get("EEC")
                    if col_eec1:
                        cell_ind = sheet.cell(row=target_row, column=col_eec1)
                        cell_ind.value = 'INDUCCION'
                        from openpyxl.styles import Font
                        cell_ind.font = Font(bold=True, color="FF0000")

        # 5. Unhide all rows in the sheet
        for r_idx in range(1, sheet.max_row + 1):
            if r_idx in sheet.row_dimensions:
                sheet.row_dimensions[r_idx].hidden = False

        # 6. Remove AutoFilter
        sheet.auto_filter.ref = None

    # 4. Save workbook is skipped in memory pipeline (coordinator script runs save).
    print("Fase 2 synchronization complete!")

    # 6. Generate PDF report for end-users
    try:
        from modules.dotacion.utils.pdf import generar_pdf
        pdf_path = os.path.splitext(cfg.OUTPUT_WORKBOOK)[0] + f"_Reporte_{year}{month:02d}.pdf"
        print(f"\n[Step 5] Generating PDF report: {pdf_path}")
        generar_pdf(
            output_path=pdf_path,
            year=year,
            month=month,
            seniority_changes=seniority_changes,
            added_advisors=added_advisors,
            status_changes=status_changes,
            supervisor_changes=supervisor_changes,
        )
    except Exception as e:
        print(f"  WARNING: PDF generation failed: {e}")
