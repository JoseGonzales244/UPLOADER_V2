import os
import datetime
import openpyxl
from typing import Optional
from modules.dotacion.dotacion_config import DotacionConfig
from modules.dotacion.utils.excel import copy_sheet_data, get_working_days

def run(wb_out, cfg: Optional[DotacionConfig] = None):
    if cfg is None:
        cfg = DotacionConfig()
    CONSOLIDADO_FILE = cfg.CONSOLIDADO_FILE
    SELECT_DOTACION_FILE = cfg.SELECT_DOTACION_FILE
    TARGET_PERIOD = cfg.TARGET_PERIOD
    
    print(f"\n--- Starting Fase 1 Monthly Cleanup for: {TARGET_PERIOD} ---")
    print(f"Consolidado file: {CONSOLIDADO_FILE}")
    print(f"Select dotacion file: {SELECT_DOTACION_FILE}")
    
    year, month = map(int, TARGET_PERIOD.split("-"))
    
    # 2. Load worksheets
    print("[Step 2] Loading dotacion sheets from consolidado and select files...")
    wb_consolidado = openpyxl.load_workbook(CONSOLIDADO_FILE, keep_links=False, data_only=True)
    
    # Handle Select Dotacion sheet character encodings
    wb_select = openpyxl.load_workbook(SELECT_DOTACION_FILE, keep_links=False, data_only=True)
    select_sheet_candidates = [n for n in wb_select.sheetnames if "SELECT" in n.upper()]
    if select_sheet_candidates:
        select_sheet_name = select_sheet_candidates[0]
    else:
        select_sheet_name = next(
            (n for n in wb_select.sheetnames if n.upper().startswith("DOTAC") or "HOJA1" in n.upper()),
            wb_select.sheetnames[0]
        )

    
    # 3. Replace DOTACIÓN and Dotación SELECT sheets
    dot_sheet_names = [n for n in wb_out.sheetnames if n.upper() in ["DOTACIÓN", "DOTACION"]]
    if dot_sheet_names:
        dot_sheet_name = dot_sheet_names[0]
    else:
        dot_sheet_name = [n for n in wb_out.sheetnames if "DOTACI" in n.upper() and "(2)" not in n and "SELECT" not in n.upper()][0]
        
    print(f"  Replacing standard DOTACION sheet: {dot_sheet_name}")
    copy_sheet_data(wb_consolidado["DOTACIÓN"], wb_out[dot_sheet_name])
    
    sel_dot_sheet_names = [n for n in wb_out.sheetnames if "SELECT" in n.upper() and "DOTACI" in n.upper()]
    if sel_dot_sheet_names:
        sel_dot_sheet_name = sel_dot_sheet_names[0]
    else:
        sel_dot_sheet_name = "Dotación SELECT"
        if sel_dot_sheet_name not in wb_out.sheetnames:
            if "DOTACIÓN (2)" in wb_out.sheetnames:
                wb_out["DOTACIÓN (2)"].title = sel_dot_sheet_name
            else:
                wb_out.create_sheet(sel_dot_sheet_name)
                
    print(f"  Replacing SELECT DOTACION sheet: {sel_dot_sheet_name}")
    copy_sheet_data(wb_select[select_sheet_name], wb_out[sel_dot_sheet_name])
    ws_sel_out = wb_out[sel_dot_sheet_name]
    for r in range(2, ws_sel_out.max_row + 1):
        for c in range(1, ws_sel_out.max_column + 1):
            val = ws_sel_out.cell(row=r, column=c).value
            if val and str(val).startswith("NOREG_"):
                ws_sel_out.cell(row=r, column=c).value = None
    
    # 4. Load holidays and calculate working days
    holidays = set()
    if "FERIADOS" in wb_consolidado.sheetnames:
        feriados_sheet = wb_consolidado["FERIADOS"]
        for row in feriados_sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:
                if isinstance(row[0], datetime.datetime):
                    holidays.add(row[0].date())
                elif isinstance(row[0], datetime.date):
                    holidays.add(row[0])
    print(f"  Loaded {len(holidays)} holidays: {holidays}")
    
    wb_consolidado.close()
    wb_select.close()
    
    working_days = get_working_days(year, month, holidays)
    print(f"  Calculated working days: {len(working_days)} days starting on {working_days[0]}")
    
    # 5. Update AVANCE DIARIO
    print("\n[Step 3] Updating AVANCE DIARIO sheet...")
    if "AVANCE DIARIO" in wb_out.sheetnames:
        av_sheet = wb_out["AVANCE DIARIO"]
        sections = [
            {"name": "Carolina", "start_row": 7},
            {"name": "Carmen", "start_row": 31},
            {"name": "Jane", "start_row": 55},
            {"name": "Karin", "start_row": 79}
        ]
        
        for sec in sections:
            print(f"  Updating section: {sec['name']} (rows {sec['start_row']} to {sec['start_row']+19})")
            for idx in range(20):
                r_idx = sec["start_row"] + idx
                date_cell = av_sheet.cell(row=r_idx, column=2)
                if idx < len(working_days):
                    date_cell.value = datetime.datetime(working_days[idx].year, working_days[idx].month, working_days[idx].day)
                else:
                    date_cell.value = None
                # Limpiar celdas C a S (columnas 3 a 19): C7:S26, C31:S50, C55:S74, C79:S98
                for col_idx in range(3, 20):
                    av_sheet.cell(row=r_idx, column=col_idx).value = None
                    
        # Fix self-referencing formulas
        for sec in sections:
            r_fix = sec["start_row"] + 19
            c_fix = 8 # Column H
            cell_val = av_sheet.cell(row=r_fix, column=c_fix).value
            if cell_val and isinstance(cell_val, str) and "'AVANCE DIARIO'!" in cell_val:
                fixed = cell_val.replace("='AVANCE DIARIO'!", "=").replace("=+'AVANCE DIARIO'!", "=")
                av_sheet.cell(row=r_fix, column=c_fix).value = fixed
                print(f"  Fixed self-ref formula at row {r_fix} col H: {cell_val} -> {fixed}")
 
    # 5.5. Clear previous monthly summary in RESULTADOS sheet
    print("\n[Step 3.5] Cleaning previous monthly summary in RESULTADOS sheet...")
    if "RESULTADOS" in wb_out.sheetnames:
        res_sheet = wb_out["RESULTADOS"]
        for r_idx in [5, 6, 7, 8]:
            # Clear Columns C to S (upper summary table)
            for col_idx in range(3, 20):
                res_sheet.cell(row=r_idx, column=col_idx).value = None
            # Clear Column V (VACACIONES / DM / OTROS)
            res_sheet.cell(row=r_idx, column=22).value = None

        # Limpiar filas manuales C18:S18, C21:S21, C24:S24, C27:S27
        for r_idx in [18, 21, 24, 27]:
            for col_idx in range(3, 20):
                res_sheet.cell(row=r_idx, column=col_idx).value = None

        print("  RESULTADOS sheet cleared (resumen superior y filas manuales 18, 21, 24, 27).")
    else:
        print("  WARNING: RESULTADOS sheet not found!")
 
    # 6. Cleanup product evaluation and EEC columns
    print("\n[Step 4] Cleaning evaluations and EEC columns in product sheets...")
    for s_name in wb_out.sheetnames:
        if s_name in ["AVANCE DIARIO", "RESULTADOS", "DOTACIÓN", "DOTACIÓN (2)", "Dotación SELECT", "Esc. Conservador (Retail+BP (2)"]:
            continue
            
        sheet = wb_out[s_name]
        headers = [sheet.cell(row=1, column=c).value for c in range(1, sheet.max_column + 1)]
        if not any(headers):
            continue
            
        eval_idx = None
        for idx, h in enumerate(headers):
            if h == "EVAL 1 (ID AUDIO FECHA)":
                eval_idx = idx + 1
                break
                
        if not eval_idx:
            continue
            
        print(f"  Cleaning sheet: {s_name}")
        eec_cols = [idx + 1 for idx, h in enumerate(headers) if h and str(h).strip().startswith("EEC")]
        col_reg_ev = headers.index("REG_EV") + 1 if "REG_EV" in headers else None
        
        last_row = 1
        if col_reg_ev:
            consec_empty = 0
            for r_idx in range(2, sheet.max_row + 1):
                reg = sheet.cell(row=r_idx, column=col_reg_ev).value
                if reg is not None and str(reg).strip() != "":
                    consec_empty = 0
                    last_row = r_idx
                else:
                    consec_empty += 1
                    if consec_empty > 20:
                        break
        else:
            consec_empty = 0
            for r_idx in range(2, sheet.max_row + 1):
                val = sheet.cell(row=r_idx, column=6).value
                if val is not None and str(val).strip() != "":
                    consec_empty = 0
                    last_row = r_idx
                else:
                    consec_empty += 1
                    if consec_empty > 20:
                        break
                        
        print(f"    Cleaning rows 2 to {last_row}")
        for r_idx in range(2, last_row + 1):
            for col_c in eec_cols:
                sheet.cell(row=r_idx, column=col_c).value = None
            for col_c in range(eval_idx, sheet.max_column + 1):
                sheet.cell(row=r_idx, column=col_c).value = None
                
    print("Fase 1 processing complete!")
