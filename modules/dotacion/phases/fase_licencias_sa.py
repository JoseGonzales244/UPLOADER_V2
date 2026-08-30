import os
import openpyxl
import pandas as pd
from copy import copy
from typing import Optional
from modules.dotacion.dotacion_config import DotacionConfig
from modules.dotacion.core.matching import mismo_supervisor


def is_backoffice(obs, puesto=None):
    """Detecta si la observación o el puesto indica puesto de BackOffice permanente (excluyendo BO interino)."""
    text_check = f"{str(obs or '')} {str(puesto or '')}".lower()
    if "interino" in text_check:
        return False
    return any(bo_kw in text_check for bo_kw in ["backoffice", "back_office", "gestion_backoffice", "asistente bo", "asistente_bo"])

def run_licencias_sa(target_period=None, licencias_file=None, consolidado_file=None, cfg: Optional[DotacionConfig] = None):
    """
    Sincroniza y genera la nueva hoja de LICENCIAS_SA para el mes solicitado.
    
    :param target_period: Cadena YYYYMM (ej. '202608').
    :param licencias_file: Ruta al archivo LICENCIAS_SA_2026.xlsx.
    :param consolidado_file: Ruta al archivo Consolidado Planilla ausentismo YYYYMM.xlsx.
    """
    if cfg is None:
        cfg = DotacionConfig(target_period or "AUTO")

    # 1. Resolver el periodo a procesar
    if target_period is None:
        target_period = f"{cfg.year}{cfg.month:02d}"
    else:
        target_period = str(target_period).strip().replace("-", "")

    target_year = int(target_period[:4])
    target_month = int(target_period[4:6])

    # Mes anterior en formato YYYYMM
    prev_m = target_month - 1 if target_month > 1 else 12
    prev_y = target_year if target_month > 1 else target_year - 1
    prev_period = f"{prev_y}{prev_m:02d}"

    # Resolver rutas
    if licencias_file is None:
        licencias_file = getattr(cfg, 'LICENCIAS_FILE', f'LICENCIAS_SA_{target_year}.xlsx')
    
    if consolidado_file is None:
        consolidado_file = getattr(cfg, 'CONSOLIDADO_FILE', f"Consolidado Planilla ausentismo {target_period}.xlsx")


    print(f"\n--- Iniciando sincronización de LICENCIAS_SA para {target_period} ---")
    print(f"Hoja base (mes anterior): {prev_period}")
    print(f"Hoja destino (mes actual): {target_period}")
    print(f"Archivo Licencias: {licencias_file}")
    print(f"Archivo Consolidado: {consolidado_file}")

    if not os.path.exists(licencias_file):
        raise FileNotFoundError(f"Archivo de licencias no encontrado: {licencias_file}")
    if not os.path.exists(consolidado_file):
        raise FileNotFoundError(f"Archivo consolidado no encontrado: {consolidado_file}")

    # 2. Cargar Workbook de Licencias
    wb_lic = openpyxl.load_workbook(licencias_file)

    if prev_period not in wb_lic.sheetnames:
        raise ValueError(f"La hoja base '{prev_period}' no se encuentra en {licencias_file}")

    ws_prev = wb_lic[prev_period]
    rows_prev = list(ws_prev.iter_rows(values_only=True))
    if len(rows_prev) < 2:
        raise ValueError(f"La hoja base '{prev_period}' está vacía o sin datos.")

    # Convertir a DataFrame la hoja del mes anterior
    headers_lic = [str(c).strip() if c is not None else f"COL_{idx}" for idx, c in enumerate(rows_prev[0])]
    df_prev = pd.DataFrame(rows_prev[1:], columns=headers_lic)
    # Filtrar filas vacías o notas al pie que no tienen REG_EJECUTIVO válido
    df_prev = df_prev[df_prev['REG_EJECUTIVO'].notna() & (df_prev['REG_EJECUTIVO'].astype(str).str.strip() != "")].copy()
    df_prev['REG_EJECUTIVO'] = df_prev['REG_EJECUTIVO'].astype(str).str.strip().str.upper()

    # Mapa de Supervisores a Equipos a partir de la hoja anterior
    sup_to_team_map = {}
    for _, r in df_prev.iterrows():
        sup = str(r['NOM_SUPERVISOR']).strip().upper() if pd.notna(r['NOM_SUPERVISOR']) else ""
        team = str(r['EQUIPO']).strip() if pd.notna(r['EQUIPO']) else ""
        if sup and team and sup not in sup_to_team_map:
            sup_to_team_map[sup] = team

    # 3. Cargar Consolidado de Planilla actual
    wb_cons = openpyxl.load_workbook(consolidado_file, data_only=True)
    dot_sheets = [s for s in wb_cons.sheetnames if s.upper().startswith("DOTACI")]
    if not dot_sheets:
        raise ValueError(f"No se encontró la hoja 'DOTACIÓN' en {consolidado_file}")
    
    ws_cons = wb_cons[dot_sheets[0]]
    rows_cons = list(ws_cons.iter_rows(values_only=True))
    headers_cons = [str(c).strip() if c is not None else "" for c in rows_cons[0]]
    df_cons = pd.DataFrame(rows_cons[1:], columns=headers_cons)

    # Filtrar dotación válida (excluyendo Cese, Provincia y BackOffice permanente)
    df_cons['REGISTRO COLABORADOR'] = df_cons['REGISTRO COLABORADOR'].astype(str).str.strip().str.upper()
    
    puesto_col = "PUESTO" if "PUESTO" in df_cons.columns else ("POSICION" if "POSICION" in df_cons.columns else None)
    
    def row_is_valid(r):
        st = str(r['ESTADO'] if 'ESTADO' in r else '').strip().upper()
        if st in ['CESADO', 'BAJA', 'INACTIVO']:
            return False
        eq = str(r['EQUIPO'] if 'EQUIPO' in r else '').strip().upper()
        if 'PROVINCIA' in eq:
            return False
        puesto_val = str(r[puesto_col] if puesto_col and puesto_col in r else '').strip().upper()
        if puesto_val in ['SUPERVISOR', 'JEFE', 'JEFA', 'GERENTE', 'SUPERVISOR/JEFE']:
            return False
        obs_val = r['OBSERVACIONES'] if 'OBSERVACIONES' in r else ""
        return not is_backoffice(obs_val, puesto_val)

    df_cons_active = df_cons[df_cons.apply(row_is_valid, axis=1)].copy()




    # Indexar Consolidado activo por REGISTRO COLABORADOR
    cons_dict = {}
    for _, r in df_cons_active.iterrows():
        reg = r['REGISTRO COLABORADOR']
        if reg and reg != 'NONE' and reg != 'NAN':
            cons_dict[reg] = r

    # 4. Procesar Ejecutivos Existentes (MANTENER LICENCIA) y actualizar Supervisor/Equipo
    new_rows = []
    processed_regs = set()

    for _, r in df_prev.iterrows():
        reg = r['REG_EJECUTIVO']
        if reg in cons_dict:
            c_row = cons_dict[reg]
            processed_regs.add(reg)

            cur_sup = str(r['NOM_SUPERVISOR']).strip() if pd.notna(r['NOM_SUPERVISOR']) else ""
            cons_sup = str(c_row['SUPERVISOR / JEFE']).strip() if pd.notna(c_row['SUPERVISOR / JEFE']) else cur_sup

            # Verificar si cambió de supervisor
            if not mismo_supervisor(cur_sup, cons_sup):
                final_sup = cons_sup
            else:
                final_sup = cur_sup

            # Determinar equipo
            final_team = sup_to_team_map.get(final_sup.upper(), r['EQUIPO'])

            new_rows.append({
                'PERIODO': int(target_period),
                'REG_EJECUTIVO': reg,
                'NOM_EJECUTIVO': r['NOM_EJECUTIVO'],
                'NOM_SUPERVISOR': final_sup,
                'EQUIPO': final_team,
                'ESTADO_LICENCIA': 'MANTENER LICENCIA'
            })

    # 5. Agregar Nuevos Ingresos (AGREGAR LICENCIA) que pertenezcan a supervisores del universo de licencias
    valid_supervisors = set(sup_to_team_map.keys())

    for reg, c_row in cons_dict.items():
        if reg not in processed_regs:
            cons_sup = str(c_row['SUPERVISOR / JEFE']).strip() if pd.notna(c_row['SUPERVISOR / JEFE']) else ""
            cons_sup_upper = cons_sup.upper()

            # Verificar si el supervisor pertenece al universo de licencias
            if any(mismo_supervisor(v_sup, cons_sup_upper) for v_sup in valid_supervisors):
                # Encontrar el equipo correspondiente
                team = 'TELEVENTAS'
                for v_sup in valid_supervisors:
                    if mismo_supervisor(v_sup, cons_sup_upper):
                        team = sup_to_team_map[v_sup]
                        break

                new_rows.append({
                    'PERIODO': int(target_period),
                    'REG_EJECUTIVO': reg,
                    'NOM_EJECUTIVO': c_row['COLABORADOR'],
                    'NOM_SUPERVISOR': cons_sup,
                    'EQUIPO': team,
                    'ESTADO_LICENCIA': 'AGREGAR LICENCIA'
                })
                processed_regs.add(reg)

    # 6. Escribir en la hoja del mes destino de forma 100% NO destructiva (preserva los IDs de relación XML)
    if target_period in wb_lic.sheetnames:
        ws_target = wb_lic[target_period]
    else:
        ws_target = wb_lic.create_sheet(title=target_period)

    headers_output = ['PERIODO', 'REG_EJECUTIVO', 'NOM_EJECUTIVO', 'NOM_SUPERVISOR', 'EQUIPO', 'ESTADO_LICENCIA']

    # Escribir encabezados en fila 1 y copiar estilos exactos del mes anterior
    for col_idx, h_val in enumerate(headers_output, 1):
        cell_tgt = ws_target.cell(row=1, column=col_idx, value=h_val)
        cell_src = ws_prev.cell(row=1, column=col_idx)
        if cell_src.has_style:
            cell_tgt.font = copy(cell_src.font)
            cell_tgt.fill = copy(cell_src.fill)
            cell_tgt.border = copy(cell_src.border)
            cell_tgt.alignment = copy(cell_src.alignment)
            cell_tgt.number_format = copy(cell_src.number_format)

    # Mes N-2 en formato YYYYMM para la fórmula de Excel
    prev_prev_m = prev_m - 1 if prev_m > 1 else 12
    prev_prev_y = prev_y if prev_m > 1 else prev_y - 1
    prev_prev_period = f"{prev_prev_y}{prev_prev_m:02d}"

    # Escribir celdas de datos, manteniendo la fórmula dinámica de Excel en ESTADO_LICENCIA
    for row_idx, item in enumerate(new_rows, 2):
        formula_val = (
            f"=IF(AND(COUNTIF('{prev_prev_period}'!B:B,B{row_idx})>0,COUNTIF('{prev_period}'!B:B,B{row_idx})>0),\"MANTENER LICENCIA\","
            f"IF(AND(COUNTIF('{prev_prev_period}'!B:B,B{row_idx})=0,COUNTIF('{prev_period}'!B:B,B{row_idx})>0),\"AGREGAR LICENCIA\","
            f"IF(AND(COUNTIF('{prev_period}'!B:B,B{row_idx})>0,COUNTIF('{target_period}'!B:B,B{row_idx})=0),\"RETIRAR LICENCIA\",\"REVISAR\")))"
        )
        row_data = [item['PERIODO'], item['REG_EJECUTIVO'], item['NOM_EJECUTIVO'], item['NOM_SUPERVISOR'], item['EQUIPO'], formula_val]
        for col_idx, val in enumerate(row_data, 1):
            cell_tgt = ws_target.cell(row=row_idx, column=col_idx, value=val)
            cell_src = ws_prev.cell(row=2, column=col_idx)
            if cell_src.has_style:
                cell_tgt.font = copy(cell_src.font)
                cell_tgt.fill = copy(cell_src.fill)
                cell_tgt.border = copy(cell_src.border)
                cell_tgt.alignment = copy(cell_src.alignment)
                cell_tgt.number_format = copy(cell_src.number_format)

    max_written_row = len(new_rows) + 1

    
    # Limpiar absolutamente todas las celdas, valores y fórmulas sobrantes de ejecuciones o pegados manuales anteriores
    old_max = max(ws_target.max_row, 300)
    for r_extra in range(max_written_row + 1, old_max + 1):
        for c_extra in range(1, 15):
            cell_tgt = ws_target.cell(row=r_extra, column=c_extra)
            cell_tgt.value = None

    new_range = f"A1:F{max_written_row}"


    # Preservar / actualizar el objeto Tabla de Excel existente (ListObject) sin modificar los IDs del ZIP
    if hasattr(ws_target, 'tables') and ws_target.tables:
        for tbl_name in list(ws_target.tables.keys()):
            tbl = ws_target.tables[tbl_name]
            tbl.ref = new_range
            if tbl.autoFilter:
                tbl.autoFilter.ref = new_range
                if hasattr(tbl.autoFilter, 'filterColumn') and tbl.autoFilter.filterColumn:
                    tbl.autoFilter.filterColumn.clear()
    else:
        # Si la hoja no tenía objeto Tabla (hoja creada de cero), añadirla oficialmente
        from openpyxl.worksheet.table import Table, TableStyleInfo, TableColumn
        table_name = f"Tabla_{target_period}"
        new_table = Table(displayName=table_name, ref=new_range)
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        new_table.tableStyleInfo = style

        for col_id, col_name in enumerate(headers_output, 1):
            new_table.tableColumns.append(TableColumn(id=col_id, name=col_name))

        ws_target.add_table(new_table)

    # Desfiltrar (desocultar) todas las filas para que todo quede visible al finalizar
    for r_dim in ws_target.row_dimensions.values():
        r_dim.hidden = False




    # Guardar libro
    wb_lic.save(licencias_file)
    wb_lic.close()

    print(f"[OK] Pestaña '{target_period}' generada exitosamente en {licencias_file}.")
    print(f"     Total registros procesados: {len(new_rows)}")
    mantenidos = sum(1 for x in new_rows if x['ESTADO_LICENCIA'] == 'MANTENER LICENCIA')
    agregados = sum(1 for x in new_rows if x['ESTADO_LICENCIA'] == 'AGREGAR LICENCIA')
    print(f"     - MANTENER LICENCIA: {mantenidos}")
    print(f"     - AGREGAR LICENCIA: {agregados}")

    return new_rows
