import os
import datetime
import openpyxl
from typing import Optional
from modules.dotacion.dotacion_config import DotacionConfig
from modules.dotacion.utils.excel import get_working_days, lock_resultados_sheet
from modules.dotacion.core.vacations import get_analyst_vacations

def run(wb, cfg: Optional[DotacionConfig] = None):
    if cfg is None:
        cfg = DotacionConfig()
    TARGET_PERIOD = cfg.TARGET_PERIOD
    print(f"\n--- Starting Fase 3 Workload Distribution for: {TARGET_PERIOD} ---")
    
    # 1. Parse holidays and calculate business days
    holidays = set()
    if "FERIADOS" in wb.sheetnames:
        feriados_sheet = wb["FERIADOS"]
        for row in feriados_sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:
                if isinstance(row[0], datetime.datetime):
                    holidays.add(row[0].date())
                elif isinstance(row[0], datetime.date):
                    holidays.add(row[0])
    year, month = map(int, TARGET_PERIOD.split("-"))
    working_days_list = get_working_days(year, month, holidays)
    num_business_days = len(working_days_list)
    print(f"  Calculated month business days: {num_business_days}")
 
    # 2. Get vacation and targets of analysts dynamically from vacation management file
    analysts = getattr(cfg, 'analysts', ["CAROLINA", "CARMEN", "JANE", "KARIN"])
    daily_targets = getattr(cfg, 'daily_targets', {
        "CAROLINA": 8,
        "CARMEN": 8,
        "JANE": 5,
        "KARIN": 8.96
    })

    vac_file = getattr(cfg, 'VACACIONES_FILE', f'Gestión de Vacaciones y Horarios {year}.xlsx')
    vacations = get_analyst_vacations(year, month, holidays, vacation_file=vac_file)
    
    capacities = {}
    assigned = {a: 0 for a in analysts}
    eval_business_days = num_business_days - 1
    for a in analysts:
        working_days = eval_business_days - vacations.get(a, 0)
        capacities[a] = working_days * daily_targets[a]
        
    print(f"  Analyst vacations: {vacations}")
    print(f"  Analyst target capacities: {capacities}")


    # 3. Identify and collect all active advisor rows in all product sheets
    sheets_to_process = ["CD", "PP", "CxC 1", "BN_B", "RCXC", "BN_C", "RTC", "PREHIP", "EC", "TC", "SEG", "SELECT"]
    active_rows = [] # list of (sheet_name, row_idx, num_evals, is_r0)
    
    # Map sheets to evals needed
    sheet_evals = {
        "TC": 1, "SEG": 2, "PP": 2, "EC": 1, "CD": 1, "PREHIP": 1,
        "SELECT": 1, "CxC 1": 2, "RTC": 1, "RCXC": 1, "BN_B": 2, "BN_C": 1
    }
    
    for s_name in sheets_to_process:
        if s_name not in wb.sheetnames:
            continue
        sheet = wb[s_name]
        headers = [sheet.cell(row=1, column=c).value for c in range(1, sheet.max_column + 1)]
        if "REG_EV" not in headers:
            continue
            
        col_reg_idx = headers.index("REG_EV") + 1 if "REG_EV" in headers else None
        col_nombre_idx = headers.index("NOMBRE_EV") + 1 if "NOMBRE_EV" in headers else (headers.index("NOMBRE") + 1 if "NOMBRE" in headers else None)
        antig_col = headers.index("ANTIGÜEDAD") + 1 if "ANTIGÜEDAD" in headers else (headers.index("ANTIGUEDAD") + 1 if "ANTIGUEDAD" in headers else None)
        eec_cols = [idx + 1 for idx, h in enumerate(headers) if h and str(h).strip().startswith("EEC")]
        col_observ_idx = headers.index("OBSERVACIONES") + 1 if "OBSERVACIONES" in headers else None
        
        consec_empty = 0
        for r_idx in range(2, sheet.max_row + 1):
            reg = sheet.cell(row=r_idx, column=col_reg_idx).value if col_reg_idx else None
            name = sheet.cell(row=r_idx, column=col_nombre_idx).value if col_nombre_idx else None
            
            is_valid_row = False
            if reg is not None and str(reg).strip() != "":
                is_valid_row = True
            elif s_name == "SELECT" and name is not None and str(name).strip() != "":
                is_valid_row = True

            if is_valid_row:
                consec_empty = 0
                reg_str = str(reg or '').strip().upper()
                name_str = str(name or '').strip().upper()
                if reg_str not in ['FECHA DEL AUDIO', 'YYYYMMDD_DNI_REGISTRO_PRODUCTO_ID_PARTE', 'CAMBIO DE TCEA - 12/06/26'] and 'B12354_TC' not in reg_str:
                    obs_val = ''
                    if col_observ_idx:
                        obs_val = str(sheet.cell(row=r_idx, column=col_observ_idx).value or '').strip().upper()
                        
                    is_backoffice = 'BACKOFFICE' in obs_val
                    # Check for hard blocking absences (LICENCIA, DM, BAJA, DESVINCULADO)
                    # Map EEC headers to get exact EEC 1 column value
                    eec_map_row = {}
                    for c_idx, h in enumerate(headers, 1):
                        if h and str(h).strip().startswith("EEC"):
                            h_clean = str(h).strip().replace(" ", "").upper()
                            eec_map_row[h_clean] = c_idx
                            
                    col_eec1_idx = eec_map_row.get("EEC1") or eec_map_row.get("EEC 1") or eec_map_row.get("EEC")
                    eec1_val = str(sheet.cell(row=r_idx, column=col_eec1_idx).value or '').strip().upper() if col_eec1_idx else ''
                        
                    is_hard_blocked = False
                    if eec_cols:
                        for col_eec in eec_cols:
                            val = str(sheet.cell(row=r_idx, column=col_eec).value or '').strip().upper()
                            if any(kw in val for kw in ['DM', 'BAJA', 'DESVINCULADO', 'VACACIONES TODO EL MES']):
                                is_hard_blocked = True
                                break
                            
                    if not is_hard_blocked and not is_backoffice:
                        antig_val = str(sheet.cell(row=r_idx, column=antig_col).value or '').strip().upper() if antig_col else ''
                        is_r0 = 'R0' in antig_val or 'INDUCCION' in eec1_val
                        is_vac = 'VACACIONES' in eec1_val and 'VACACIONES TODO' not in eec1_val
                        
                        base_evals = sheet_evals.get(s_name, 1)
                        
                        if (is_r0 and s_name != "BN_B") or is_vac:
                            # In 2-eval products, if EEC 1 has INDUCCION or VACACIONES, assign 1 eval to EEC 2!
                            evals_needed = 1 if base_evals == 2 else 0
                            only_eec2 = True if base_evals == 2 else False
                        else:
                            evals_needed = base_evals
                            only_eec2 = False
                            
                        if evals_needed > 0:
                            active_rows.append((s_name, r_idx, evals_needed, only_eec2))
            else:
                consec_empty += 1
                if consec_empty > 20:
                    break
                    
    total_evals_to_distribute = sum(r[2] for r in active_rows)
    print(f"  Total evaluations to distribute in workbook: {total_evals_to_distribute}")
    
    # 4. Proportional distribution calculation
    total_capacity = sum(capacities.values())
    if total_capacity == 0:
        raise ValueError("Total capacity of analysts is zero! Check vacations configuration.")
        
    adjusted_targets = {}
    for a in analysts:
        ratio = capacities[a] / total_capacity
        adjusted_targets[a] = round(ratio * total_evals_to_distribute)
        
    # Reconcile rounding discrepancies
    discrepancy = total_evals_to_distribute - sum(adjusted_targets.values())
    if discrepancy != 0:
        # Distribute discrepancy to analyst with largest capacity
        sorted_analysts = sorted(analysts, key=lambda x: capacities[x], reverse=True)
        adjusted_targets[sorted_analysts[0]] += discrepancy
        
    print(f"  Adjusted workload targets: {adjusted_targets}")
    
    # 5. Distribute workloads to advisors
    # Prioritizar SELECT primero para que Karin absorba preferentemente su cuota en SELECT
    active_rows.sort(key=lambda x: (0 if x[0] == "SELECT" else 1, -x[2], x[0]))
    
    distribution_log = [] # list of (analyst, sheet, row, count, only_eec2)
    
    for s_name, r_idx, evals, only_eec2 in active_rows:
        # Find analyst with the lowest ratio of assigned / adjusted_target (or absolute difference)
        best_analyst = None
        best_score = float('inf')
        
        for a in analysts:
            tgt = adjusted_targets[a]
            if tgt == 0:
                continue
            # Score: how far they are from their target proportional workload
            score = (assigned[a] + evals) / tgt
            # Preferencia para KARIN en la hoja SELECT mientras no exceda su target
            if s_name == "SELECT" and a == "KARIN" and (assigned[a] + evals) <= tgt:
                score -= 0.35
            if score < best_score:
                best_score = score
                best_analyst = a
                
        if not best_analyst:
            # Fallback to analyst with highest capacity
            best_analyst = sorted(analysts, key=lambda x: capacities[x], reverse=True)[0]
            
        assigned[best_analyst] += evals
        distribution_log.append((best_analyst, s_name, r_idx, evals, only_eec2))
        
    print(f"  Distribution complete. Assigned counts: {assigned}")

    # 6. Write assignments back to the product sheets (using BOLD BLACK font)
    from openpyxl.styles import Font
    bold_black_font = Font(color="000000", bold=True)

    for analyst, s_name, r_idx, evals, only_eec2 in distribution_log:
        sheet = wb[s_name]
        headers = [sheet.cell(row=1, column=c).value for c in range(1, sheet.max_column + 1)]
        eec_map = {}
        for c_idx, h in enumerate(headers, 1):
            if h and str(h).strip().startswith("EEC"):
                h_clean = str(h).strip().replace(" ", "").upper()
                eec_map[h_clean] = c_idx
                
        if only_eec2:
            # Assign ONLY to EEC 2 (slot 2), while EEC 1 retains INDUCCION or VACACIONES
            slot_keys = ["EEC2", "EEC 2"]
            col_idx = None
            for sk in slot_keys:
                if sk in eec_map:
                    col_idx = eec_map[sk]
                    break
            if col_idx:
                cell = sheet.cell(row=r_idx, column=col_idx, value=analyst)
                cell.font = bold_black_font
        else:
            # Regular R3 / R2 / R1 advisor
            for slot in range(1, evals + 1):
                slot_keys = [f"EEC{slot}", f"EEC {slot}"]
                if slot == 1:
                    slot_keys.append("EEC")
                    
                col_idx = None
                for sk in slot_keys:
                    if sk in eec_map:
                        col_idx = eec_map[sk]
                        break
                        
                if col_idx:
                    cell = sheet.cell(row=r_idx, column=col_idx, value=analyst)
                    cell.font = bold_black_font

    # 7. Update RESULTADOS sheet summary
    print("[Step 3] Updating RESULTADOS sheet...")
    if "RESULTADOS" in wb.sheetnames:
        res_sheet = wb["RESULTADOS"]
        
        # Column mapping for RESULTADOS: col_idx -> (list_of_sheet_names, eval_slot)
        col_sheets_map = {
            3: (['TC'], 1),
            4: (['TC'], 2),
            5: (['SEG'], 1),
            6: (['SEG'], 2),
            7: (['PP'], 1),
            8: (['PP'], 2),
            9: (['EC'], 1),
            10: (['CD'], 1),
            11: (['PREHIP'], 1),
            12: (['SELECT'], 1),
            13: (['CxC 1'], 1),
            14: (['CxC 1'], 2),
            15: (['RTC'], 1),
            16: (['RCXC'], 1),
            17: (['BN_B'], 1),
            18: (['BN_B'], 2),
            19: (['BN_C'], 1)
        }
        
        res_analysts = ["CAROLINA", "CARMEN", "JANE", "KARIN"]
        
        for idx, a in enumerate(res_analysts):
            row_idx = 5 + idx
            eval_realiz_row = 17 + idx * 3
            
            for col_idx, (s_list, eval_slot) in col_sheets_map.items():
                count = 0
                for s_name in s_list:
                    if s_name in wb.sheetnames:
                        sheet = wb[s_name]
                        headers = [sheet.cell(row=1, column=c).value for c in range(1, sheet.max_column + 1)]
                        eec_map = {}
                        for c_i, h in enumerate(headers, 1):
                            if h and str(h).strip().startswith("EEC"):
                                h_clean = str(h).strip().replace(" ", "").upper()
                                eec_map[h_clean] = c_i
                                
                        slot_keys = [f"EEC{eval_slot}", f"EEC {eval_slot}"]
                        if eval_slot == 1:
                            slot_keys.append("EEC")
                            
                        col_c = None
                        for sk in slot_keys:
                            if sk in eec_map:
                                col_c = eec_map[sk]
                                break
                                
                        if col_c:
                            for r_i in range(2, sheet.max_row + 1):
                                val = str(sheet.cell(row=r_i, column=col_c).value or "").strip().upper()
                                if val == a:
                                    count += 1
                                    
                val_to_write = count if count > 0 else None
                # Write to upper summary table (rows 5..8) ONLY
                res_sheet.cell(row=row_idx, column=col_idx, value=val_to_write)
                # Note: Lower table rows (17, 20, 23, 26) retain their original formulas referencing AVANCE DIARIO!
                        
            # Write active working days to Column U (21) and vacation days to Column V (22)
            vac_val = vacations[a]
            base_eval_days = eval_business_days
            analyst_working_days = base_eval_days - vac_val
            res_sheet.cell(row=row_idx, column=21, value=analyst_working_days)
            res_sheet.cell(row=row_idx, column=22, value=vac_val if vac_val > 0 else None)

        # Re-lock RESULTADOS sheet protection
        lock_resultados_sheet(wb)
    else:
        print("  WARNING: RESULTADOS sheet not found!")

    print("Fase 3 workload distribution complete!")
