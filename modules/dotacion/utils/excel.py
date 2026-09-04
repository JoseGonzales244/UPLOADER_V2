import datetime
import openpyxl
from copy import copy

def find_headers_and_row(sheet, search_term):
    search_terms = [search_term] if isinstance(search_term, str) else list(search_term)
    for r_idx in range(1, 15):
        row_vals = [sheet.cell(row=r_idx, column=c).value for c in range(1, sheet.max_column + 1)]
        if any(row_vals):
            for term in search_terms:
                term_upper = str(term).strip().upper()
                if any(term_upper in str(v).strip().upper() for v in row_vals if v is not None):
                    headers = [str(v).strip() if v is not None else None for v in row_vals]
                    return headers, r_idx
    raise ValueError(f"Could not find header row containing any of {search_terms} in sheet '{sheet.title}'")

def find_advisor_reg_col_idx(headers):
    """Encuentra el índice de la columna de registro del asesor (evitando columnas de supervisor o jefe)."""
    candidates = ["REGISTRO COLABORADOR", "REG_COLAB", "REG_PROMOTOR", "REG_EJECUTIVO", "REGISTRO"]
    for cand in candidates:
        for idx, h in enumerate(headers):
            if h and str(h).strip().upper() == cand:
                return idx
    for idx, h in enumerate(headers):
        if h:
            h_upper = str(h).strip().upper()
            if any(c in h_upper for c in candidates) and "SUPER" not in h_upper and "JEFE" not in h_upper:
                return idx
    return None

def copy_row_style(ws, src_row, tgt_row):
    for col_idx in range(1, ws.max_column + 1):
        src_cell = ws.cell(row=src_row, column=col_idx)
        tgt_cell = ws.cell(row=tgt_row, column=col_idx)
        if src_cell.has_style:
            tgt_cell.font = copy(src_cell.font)
            tgt_cell.border = copy(src_cell.border)
            tgt_cell.fill = copy(src_cell.fill)
            tgt_cell.number_format = copy(src_cell.number_format)
            tgt_cell.alignment = copy(src_cell.alignment)

def copy_sheet_data(source_ws, target_ws, period_filter=None):
    # Encontrar la fila de cabecera dinámicamente en el origen
    headers = None
    hdr_row = 1
    reg_candidates = ["REGISTRO COLABORADOR", "REG_COLAB", "REG_PROMOTOR", "REG_EJECUTIVO", "REGISTRO"]
    for r in range(1, 15):
        vals = [source_ws.cell(row=r, column=c).value for c in range(1, source_ws.max_column+1)]
        if any(vals) and any(any(cand in str(v).upper() for cand in reg_candidates) for v in vals if v):
            headers = [str(v).strip() if v is not None else "" for v in vals]
            hdr_row = r
            break
            
    if not headers:
        # Fallback simple
        headers = [str(source_ws.cell(row=1, column=c).value or "") for c in range(1, source_ws.max_column+1)]
        hdr_row = 1

    period_col_idx = None
    if period_filter and "PERIODO" in headers:
        period_col_idx = headers.index("PERIODO") + 1

    # Limpiar destino sin alterar estructura ni causar corrupción de celdas combinadas
    for r in range(1, target_ws.max_row + 1):
        for c in range(1, target_ws.max_column + 1):
            target_ws.cell(row=r, column=c).value = None

    # Normalizar cabecera de registro en destino si viene como REG_COLAB, REG_PROMOTOR o REG_EJECUTIVO
    norm_headers = list(headers)
    for col_idx, h_val in enumerate(norm_headers):
        h_upper = str(h_val).strip().upper()
        if h_upper in ["REG_COLAB", "REG_PROMOTOR", "REG_EJECUTIVO"]:
            norm_headers[col_idx] = "REGISTRO"

    # Copiar cabecera
    for col_idx, h_val in enumerate(norm_headers):
        target_ws.cell(row=1, column=col_idx + 1, value=h_val)

    # Copiar datos
    tgt_row_idx = 2
    for src_row_idx in range(hdr_row + 1, source_ws.max_row + 1):
        if period_col_idx and period_filter:
            p_val = source_ws.cell(row=src_row_idx, column=period_col_idx).value
            if p_val and str(p_val).strip() != str(period_filter).strip():
                continue # Saltar fila si no coincide el periodo
                
        # Copiar todos los valores de la fila
        colab_val = None
        reg_col_idx = None
        for col_idx in range(1, len(headers) + 1):
            h_name = str(headers[col_idx-1]).upper()
            if "COLABORADOR" in h_name or "NOMBRE" in h_name:
                colab_val = source_ws.cell(row=src_row_idx, column=col_idx).value
            if any(cand == h_name.strip() for cand in reg_candidates) and "SUPER" not in h_name and "JEFE" not in h_name:
                reg_col_idx = col_idx

        for col_idx in range(1, len(headers) + 1):
            val = source_ws.cell(row=src_row_idx, column=col_idx).value
            if col_idx == reg_col_idx and (val is None or str(val).strip() == "") and colab_val:
                val = f"NOREG_{src_row_idx}"
            target_ws.cell(row=tgt_row_idx, column=col_idx, value=val)
        tgt_row_idx += 1

def get_working_days(year, month, holidays_set):
    # Encontrar el 4to día útil del mes de evaluación
    count = 0
    curr_date = datetime.date(year, month, 1)
    start_working_day = None
    while curr_date.month == month:
        if curr_date.weekday() < 5 and curr_date not in holidays_set:
            count += 1
            if count == 4:
                start_working_day = curr_date
                break
        curr_date += datetime.timedelta(days=1)
        
    if not start_working_day:
        raise ValueError(f"Could not find the 4th business day of month {month} in year {year}.")
        
    working_days = []
    curr_date = start_working_day
    while curr_date.month == month:
        if curr_date.weekday() < 5 and curr_date not in holidays_set:
            working_days.append(curr_date)
        curr_date += datetime.timedelta(days=1)
        
    return working_days

def recreate_unlocked_resultados_sheet(wb):
    if "RESULTADOS" not in wb.sheetnames:
        return
    ws = wb["RESULTADOS"]
    ws.protection.sheet = False
    ws.protection.disable()
    unlocked_protection = openpyxl.styles.Protection(locked=False)
    for row in ws.iter_rows():
        for cell in row:
            cell.protection = copy(unlocked_protection)
def lock_resultados_sheet(wb):
    if "RESULTADOS" not in wb.sheetnames:
        return
    ws = wb["RESULTADOS"]
    ws.protection.sheet = True
    ws.protection.objects = True
    ws.protection.scenarios = True
    print("  [OK] RESULTADOS sheet protection restored (re-locked)!")


def clean_broken_defined_names(wb):
    """Elimina rangos con nombre corruptos (#REF!) y vistas personalizadas (Z_...) que provocan avisos en Excel."""
    # 1. Global defined names
    global_del = []
    for name in list(wb.defined_names):
        try:
            dn = wb.defined_names[name]
            val = getattr(dn, 'attr_text', '') or getattr(dn, 'value', '') or str(dn)
            if '#REF!' in str(val) or str(name).startswith('Z_'):
                global_del.append(name)
        except Exception:
            pass

    for k in global_del:
        try:
            del wb.defined_names[k]
        except Exception:
            pass

    # 2. Sheet-level defined names
    total_sheet_del = 0
    for ws in wb.worksheets:
        sheet_del = []
        for name in list(getattr(ws, 'defined_names', [])):
            if name.startswith('Z_') or 'wvu' in name or 'FilterData' in name:
                sheet_del.append(name)
        for k in sheet_del:
            try:
                del ws.defined_names[k]
                total_sheet_del += 1
            except Exception:
                pass

    if global_del or total_sheet_del:
        print(f"  [OK] Limpiados {len(global_del)} rangos globales y {total_sheet_del} de hoja corruptos/Z_.")


def cleanup_phantom_rows(wb):
    """No-op seguro: la manipulación de ws._cells corrompe las dimensiones internas OpenXML."""
    pass


def restore_pristine_xml_structures(orig_path, target_path):
    """No-op seguro: la inyección de calcChain/printerSettings sin declarar en Content_Types corrompe el paquete OpenXML."""
    pass





